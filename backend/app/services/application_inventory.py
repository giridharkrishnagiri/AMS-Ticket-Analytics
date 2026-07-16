from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from dataclasses import dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import and_, func, select, text, update
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models import ApplicationInventoryItem, AssessmentOutOfScopeTicket, Project, Ticket
from app.services.application_metrics import recompute_application_ticket_user_metrics
from app.services.ingestion import (
    CSV_ENCODING_CANDIDATES,
    INGESTION_BATCH_SIZE,
    ParsedSourceRow,
    build_raw_data,
    detect_csv_encoding,
    make_unique_headers,
    normalize_source_column_name,
    row_has_any_value,
)
from app.services.mapping import parse_bool_value, text_or_none
from app.services.sap_classification import derive_sap_non_sap

MAX_MESSAGE_SAMPLES = 50
TOP_UNMATCHED_LIMIT = 25
UNMATCHED_SAMPLE_LIMIT = 5
TARGET_WORKSHEET_NAME = "Group-App-BizService"
TARGET_WORKSHEET_NAMES = ("Group-App-BizService-Inscope", "Group-App-BizService")
CMDB_BLANK_MARKERS = {"#N/A"}

CORE_FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "application_number_apm": (
        "Application Number (APM)",
        "Application Number",
        "APM",
        "App Number",
        "Application ID",
        "application_number_apm",
    ),
    "parent_application_name": (
        "Parent Business Application",
        "Parent Business App",
        "Parent Application",
        "Parent Application Name",
        "parent_business_application",
    ),
    "assignment_group": (
        "Support group",
        "Support Group",
        "Support group name",
        "Support Group Name",
        "Assignment Group",
        "Assignment group",
        "Assignment Groups",
        "Assigment Groups",
        "assignment_group",
        "Managed By Group",
        "Managed by group",
    ),
    "assignment_group_owner": (
        "Support group's owner",
        "Support group owner",
        "Assignment Group Owner",
    ),
    "application_owner": (
        "Owned by",
        "Owned By",
        "Application Owner",
        "Application owner",
        "Owner",
        "Business Owner",
    ),
    "business_service_ci_name": (
        "Business Service CI Name",
        "Business Service CI",
        "Business Service",
        "Business_Service",
        "Name",
        "Application",
        "Application Name",
        "business_service_ci_name",
    ),
    "support_lead": (
        "Support Lead (Managed by)",
        "Support Lead",
        "Managed by",
        "Managed By",
        "Manager",
        "Group Manager",
        "Assignment Group Manager",
    ),
    "functional_track": (
        "Functional Track",
        "Functional track",
        "functional_track",
        "Track",
        "AMS Functional Track",
        "Portfolio Track",
    ),
    "ams_owner": ("AMS Lead", "AMS Owner", "AMS owner", "ams_owner", "Application AMS Owner"),
    "supported_by_vendor": (
        "Supported By Vendor",
        "Supported by Vendor",
        "Supported By",
        "Support vendor",
        "Vendor",
    ),
    "service_type": ("Service Type", "Service type", "service_type", "ServiceType"),
    "service_entitlement": (
        "Service Entitlement",
        "Service entitlement",
        "service_entitlement",
        "ServiceEntitlement",
    ),
    "sap_non_sap": (
        "Subcategory",
        "SAP / Non-SAP",
        "SAP Non-SAP",
        "SAP_NonSAP",
        "Application Subcategory",
    ),
    "hosting_env": (
        "Hosting Env",
        "Hosting Environment",
        "hosting_env",
        "Used for",
        "Used For",
        "Environment",
        "Usage",
    ),
    "global_application": ("Global", "Global Application", "global_application"),
    "scope_status": (
        "In Scope / Out of Scope",
        "Scope",
        "Scope Status",
        "Application Scope",
        "In Scope",
        "In-Scope",
        "InScope",
        "AMS Scope",
    ),
    "lifecycle_stage_status": (
        "Life Cycle Stage Status",
        "Lifecycle Stage Status",
        "lifecycle_stage_status",
    ),
    "lifecycle_current": (
        "Lifecycle - Current",
        "Lifecycle Current",
        "lifecycle_current",
    ),
    "lifecycle_1_to_3_years": (
        "Lifecycle - 1 to 3 years",
        "Lifecycle 1 to 3 years",
        "Lifecycle - 1-3 years",
        "lifecycle_1_to_3_years",
    ),
    "lifecycle_3_to_5_years": (
        "Lifecycle - 3 to 5 years",
        "Lifecycle 3 to 5 years",
        "Lifecycle - 3-5 years",
        "lifecycle_3_to_5_years",
    ),
    "active": ("Active",),
    "active_users": ("Active Users", "Active User", "Active Users Count"),
}
CORE_FIELD_NAMES = set(CORE_FIELD_ALIASES)
CORE_ALIAS_TO_FIELD = {
    normalize_source_column_name(alias): field_name
    for field_name, aliases in CORE_FIELD_ALIASES.items()
    for alias in aliases
}

AMS_OWNER_FUNCTIONAL_TRACK_FALLBACKS = {
    "ashwin rao": "Digital Finance",
    "luboslav matisko": "Enterprise Apps & RPA",
    "luis sanchez": "S/4 T2S / NALA",
    "matyas hunyady": "Supply Chain",
    "ravi telu": "Marketing, Sales & eCommerce",
    "seshu avala": "Data & Analytics",
}


class ApplicationInventoryError(Exception):
    pass


@dataclass(frozen=True)
class ValueCount:
    value: str
    count: int


@dataclass(frozen=True)
class UnmatchedBusinessService:
    business_service: str
    ticket_count: int
    assignment_group_count: int
    sample_assignment_groups: list[str]
    sample_ticket_numbers: list[str]


@dataclass
class InventoryUploadResult:
    project_id: UUID
    total_rows: int = 0
    inserted_count: int = 0
    updated_count: int = 0
    skipped_count: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    distinct_business_services: set[str] = field(default_factory=set)
    distinct_parent_applications: set[str] = field(default_factory=set)
    distinct_assignment_groups: set[str] = field(default_factory=set)
    distinct_application_owners: set[str] = field(default_factory=set)
    distinct_support_leads: set[str] = field(default_factory=set)
    distinct_functional_tracks: set[str] = field(default_factory=set)
    distinct_ams_owners: set[str] = field(default_factory=set)
    distinct_supported_vendors: set[str] = field(default_factory=set)
    distinct_hosting_envs: set[str] = field(default_factory=set)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass
class InventoryActiveUsersUpdateResult:
    project_id: UUID
    total_rows: int = 0
    matched_count: int = 0
    updated_count: int = 0
    unchanged_count: int = 0
    unmatched_count: int = 0
    skipped_count: int = 0
    invalid_count: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass
class InventoryHostingEnvUpdateResult:
    project_id: UUID
    total_rows: int = 0
    matched_count: int = 0
    updated_count: int = 0
    unchanged_count: int = 0
    unmatched_count: int = 0
    skipped_count: int = 0
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass(frozen=True)
class InventoryEnrichmentSummary:
    project_id: UUID
    total_tickets: int
    matched_tickets: int
    unmatched_tickets: int
    updated_tickets: int
    match_rate_pct: float | None
    matched_by_business_service_count: int
    matched_by_application_count: int
    unmatched_business_service_count: int
    distinct_ticket_business_service_count: int
    distinct_inventory_business_service_count: int
    top_unmatched_business_services: list[ValueCount]
    top_unmatched_applications: list[ValueCount]
    top_unmatched_assignment_groups: list[ValueCount]


@dataclass(frozen=True)
class BusinessServiceCoverage:
    project_id: UUID
    distinct_ticket_business_service_count: int
    distinct_inventory_business_service_count: int
    matched_business_service_count: int
    unmatched_business_service_count: int
    business_service_coverage_pct: float | None
    rows: list[UnmatchedBusinessService]


@dataclass(frozen=True)
class ScopeSummary:
    project_id: UUID
    in_scope_tickets: int
    out_of_scope_tickets: int
    total_classified_tickets: int
    in_scope_pct: float | None
    out_of_scope_pct: float | None
    distinct_in_scope_assignment_groups: int
    distinct_out_of_scope_assignment_groups: int
    top_out_of_scope_assignment_groups: list[ValueCount]
    top_out_of_scope_business_services: list[ValueCount]


def append_sample_message(messages: list[str], message: str) -> None:
    if len(messages) < MAX_MESSAGE_SAMPLES:
        messages.append(message)


def normalize_match_key(value: Any) -> str | None:
    text = text_or_none(value)
    if text is None:
        return None
    normalized = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip().casefold()
    return normalized or None


def get_raw_value(raw_data: dict[str, Any], field_name: str) -> Any:
    for alias in CORE_FIELD_ALIASES[field_name]:
        if alias in raw_data:
            return raw_data[alias]

    normalized_lookup = {
        normalize_source_column_name(column_name): value
        for column_name, value in raw_data.items()
    }
    for alias in CORE_FIELD_ALIASES[field_name]:
        normalized_alias = normalize_source_column_name(alias)
        if normalized_alias in normalized_lookup:
            return normalized_lookup[normalized_alias]
    return None


def raw_data_has_core_field(raw_data: dict[str, Any], field_name: str) -> bool:
    normalized_headers = {
        normalize_source_column_name(column_name) for column_name in raw_data
    }
    return any(
        normalize_source_column_name(alias) in normalized_headers
        for alias in CORE_FIELD_ALIASES[field_name]
    )


def build_cmdb_payload(raw_data: dict[str, Any]) -> dict[str, Any]:
    payload: dict[str, Any] = {}
    for column_name, value in raw_data.items():
        normalized_name = normalize_source_column_name(column_name)
        if CORE_ALIAS_TO_FIELD.get(normalized_name) in CORE_FIELD_NAMES:
            continue
        payload[column_name] = cmdb_payload_value_or_none(value)
    return payload


def cmdb_payload_value_or_none(value: Any) -> Any:
    if isinstance(value, str):
        text = text_or_none(value)
        if text is None or text.upper() in CMDB_BLANK_MARKERS:
            return None
        return text
    return value


def parse_active(value: Any, row_number: int, result: InventoryUploadResult) -> bool | None:
    text = text_or_none(value)
    if text is None:
        return True

    parsed = parse_bool_value(value)
    if parsed is None:
        append_sample_message(
            result.warnings,
            f"Row {row_number}: Active value '{text}' could not be parsed.",
        )
    return parsed


def parse_active_users(value: Any, row_number: int, result: InventoryUploadResult) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        append_sample_message(
            result.warnings,
            f"Row {row_number}: Active Users value '{value}' could not be parsed.",
        )
        return None
    if isinstance(value, int):
        if value >= 0:
            return value
        append_sample_message(
            result.warnings,
            f"Row {row_number}: Active Users value '{value}' could not be parsed.",
        )
        return None
    if isinstance(value, float):
        if value >= 0 and value.is_integer():
            return int(value)
        append_sample_message(
            result.warnings,
            f"Row {row_number}: Active Users value '{value}' could not be parsed.",
        )
        return None

    text = text_or_none(value)
    if text is None:
        return None
    normalized = text.replace(",", "")
    try:
        parsed = float(normalized)
    except ValueError:
        append_sample_message(
            result.warnings,
            f"Row {row_number}: Active Users value '{text}' could not be parsed.",
        )
        return None
    if parsed >= 0 and parsed.is_integer():
        return int(parsed)
    append_sample_message(
        result.warnings,
        f"Row {row_number}: Active Users value '{text}' could not be parsed.",
    )
    return None


def normalize_scope_status(value: Any) -> str:
    text = text_or_none(value)
    if text is None:
        return "unknown"
    normalized = " ".join(text.replace("-", " ").split()).casefold()
    if normalized in {"in scope", "inscope", "yes", "y", "true"}:
        return "in_scope"
    if normalized in {"out of scope", "out scope", "outofscope", "no", "n", "false"}:
        return "out_of_scope"
    return "unknown"


def derive_inventory_scope_status(raw_scope: Any) -> str:
    explicit_scope = text_or_none(raw_scope)
    if explicit_scope is not None:
        scope_status = normalize_scope_status(explicit_scope)
        return scope_status
    return "unknown"


def functional_track_from_ams_owner(ams_owner: Any) -> str | None:
    ams_owner_key = normalize_match_key(ams_owner)
    if ams_owner_key is None:
        return None
    return AMS_OWNER_FUNCTIONAL_TRACK_FALLBACKS.get(ams_owner_key)


def clean_inventory_values(
    project_id: UUID,
    source_filename: str,
    parsed_row: ParsedSourceRow,
    result: InventoryUploadResult,
) -> dict[str, Any] | None:
    assignment_group = text_or_none(get_raw_value(parsed_row.raw_data, "assignment_group"))
    business_service_ci_name = text_or_none(
        get_raw_value(parsed_row.raw_data, "business_service_ci_name")
    )
    ams_owner = text_or_none(get_raw_value(parsed_row.raw_data, "ams_owner"))
    functional_track = text_or_none(get_raw_value(parsed_row.raw_data, "functional_track"))
    values: dict[str, Any] = {
        "project_id": project_id,
        "application_number_apm": text_or_none(
            get_raw_value(parsed_row.raw_data, "application_number_apm")
        ),
        "parent_application_name": text_or_none(
            get_raw_value(parsed_row.raw_data, "parent_application_name")
        ),
        "assignment_group": assignment_group,
        "assignment_group_owner": text_or_none(
            get_raw_value(parsed_row.raw_data, "assignment_group_owner")
        ),
        "application_owner": text_or_none(
            get_raw_value(parsed_row.raw_data, "application_owner")
        ),
        "business_service_ci_name": business_service_ci_name or "",
        "support_lead": text_or_none(get_raw_value(parsed_row.raw_data, "support_lead")),
        "functional_track": functional_track,
        "ams_owner": ams_owner,
        "supported_by_vendor": text_or_none(
            get_raw_value(parsed_row.raw_data, "supported_by_vendor")
        ),
        "service_type": text_or_none(get_raw_value(parsed_row.raw_data, "service_type")),
        "service_entitlement": text_or_none(
            get_raw_value(parsed_row.raw_data, "service_entitlement")
        ),
        "hosting_env": text_or_none(get_raw_value(parsed_row.raw_data, "hosting_env")),
        "global_application": text_or_none(
            get_raw_value(parsed_row.raw_data, "global_application")
        ),
        "scope_status": "unknown",
        "lifecycle_stage_status": text_or_none(
            get_raw_value(parsed_row.raw_data, "lifecycle_stage_status")
        ),
        "lifecycle_current": text_or_none(
            get_raw_value(parsed_row.raw_data, "lifecycle_current")
        ),
        "lifecycle_1_to_3_years": text_or_none(
            get_raw_value(parsed_row.raw_data, "lifecycle_1_to_3_years")
        ),
        "lifecycle_3_to_5_years": text_or_none(
            get_raw_value(parsed_row.raw_data, "lifecycle_3_to_5_years")
        ),
        "sap_non_sap": text_or_none(get_raw_value(parsed_row.raw_data, "sap_non_sap"))
        or derive_sap_non_sap(assignment_group),
        "active": parse_active(
            get_raw_value(parsed_row.raw_data, "active"),
            parsed_row.row_number,
            result,
        ),
        "active_users": parse_active_users(
            get_raw_value(parsed_row.raw_data, "active_users"),
            parsed_row.row_number,
            result,
        ),
        "cmdb_payload": build_cmdb_payload(parsed_row.raw_data),
        "source_filename": source_filename,
        "source_sheet_name": parsed_row.source_sheet_name,
        "is_current": True,
        "replaced_at": None,
        "source_row_number": parsed_row.row_number,
    }

    if assignment_group is None and business_service_ci_name is None:
        result.skipped_count += 1
        append_sample_message(
            result.errors,
            f"Row {parsed_row.row_number}: Support group name or Business Service CI Name "
            "is required.",
        )
        return None

    track_upload_distincts(result, values)
    return values


def track_upload_distincts(result: InventoryUploadResult, values: dict[str, Any]) -> None:
    distinct_fields = (
        ("business_service_ci_name", result.distinct_business_services),
        ("parent_application_name", result.distinct_parent_applications),
        ("assignment_group", result.distinct_assignment_groups),
        ("application_owner", result.distinct_application_owners),
        ("support_lead", result.distinct_support_leads),
        ("functional_track", result.distinct_functional_tracks),
        ("ams_owner", result.distinct_ams_owners),
        ("supported_by_vendor", result.distinct_supported_vendors),
        ("hosting_env", result.distinct_hosting_envs),
    )
    for field_name, target_set in distinct_fields:
        value = normalize_match_key(values.get(field_name))
        if value is not None:
            target_set.add(value)


def duplicate_key_condition(values: dict[str, Any]) -> Any:
    return and_(
        ApplicationInventoryItem.project_id == values["project_id"],
        ApplicationInventoryItem.is_current.is_(True),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.business_service_ci_name)),
            "",
        )
        == (normalize_match_key(values["business_service_ci_name"]) or ""),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.parent_application_name)),
            "",
        )
        == (normalize_match_key(values["parent_application_name"]) or ""),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.assignment_group)),
            "",
        )
        == (normalize_match_key(values["assignment_group"]) or ""),
    )


def inventory_match_condition(
    project_id: UUID,
    *,
    business_service_ci_name: str,
    parent_application_name: str | None,
    assignment_group: str | None,
) -> Any:
    return and_(
        ApplicationInventoryItem.project_id == project_id,
        ApplicationInventoryItem.is_current.is_(True),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.business_service_ci_name)),
            "",
        )
        == (normalize_match_key(business_service_ci_name) or ""),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.parent_application_name)),
            "",
        )
        == (normalize_match_key(parent_application_name) or ""),
        func.coalesce(
            func.lower(func.btrim(ApplicationInventoryItem.assignment_group)),
            "",
        )
        == (normalize_match_key(assignment_group) or ""),
    )


def upsert_inventory_item(db: Session, values: dict[str, Any]) -> bool:
    item = db.scalar(
        select(ApplicationInventoryItem).where(duplicate_key_condition(values)).limit(1)
    )
    if item is None:
        db.add(ApplicationInventoryItem(**values))
        db.flush()
        return True

    for field_name, value in values.items():
        setattr(item, field_name, value)
    db.flush()
    return False


def inventory_values_key(values: dict[str, Any]) -> tuple[str, str, str]:
    return (
        normalize_match_key(values["business_service_ci_name"]) or "",
        normalize_match_key(values["parent_application_name"]) or "",
        normalize_match_key(values["assignment_group"]) or "",
    )


def update_application_inventory_active_users_from_file(
    db: Session,
    project_id: UUID,
    path: Path,
) -> InventoryActiveUsersUpdateResult:
    """Update only active_users from an inventory workbook; never insert or upsert rows."""
    ensure_project_exists(db, project_id)
    result = InventoryActiveUsersUpdateResult(project_id=project_id)

    try:
        for parsed_row in iter_inventory_file_rows(path, result):
            result.total_rows += 1
            if not raw_data_has_core_field(parsed_row.raw_data, "active_users"):
                raise ApplicationInventoryError("Active Users column was not found.")

            business_service_ci_name = text_or_none(
                get_raw_value(parsed_row.raw_data, "business_service_ci_name")
            )
            if business_service_ci_name is None:
                result.skipped_count += 1
                append_sample_message(
                    result.errors,
                    f"Row {parsed_row.row_number}: Business Service CI Name is required.",
                )
                continue

            raw_active_users = get_raw_value(parsed_row.raw_data, "active_users")
            raw_active_users_text = text_or_none(raw_active_users)
            warning_count_before = result.warning_count
            active_users = parse_active_users(
                raw_active_users,
                parsed_row.row_number,
                result,
            )
            if (
                raw_active_users_text is not None
                and active_users is None
                and result.warning_count > warning_count_before
            ):
                result.invalid_count += 1
                result.skipped_count += 1
                continue

            parent_application_name = text_or_none(
                get_raw_value(parsed_row.raw_data, "parent_application_name")
            )
            assignment_group = text_or_none(get_raw_value(parsed_row.raw_data, "assignment_group"))
            item = db.scalar(
                select(ApplicationInventoryItem)
                .where(
                    inventory_match_condition(
                        project_id,
                        business_service_ci_name=business_service_ci_name,
                        parent_application_name=parent_application_name,
                        assignment_group=assignment_group,
                    )
                )
                .limit(1)
            )
            if item is None:
                result.unmatched_count += 1
                append_sample_message(
                    result.warnings,
                    "Row "
                    f"{parsed_row.row_number}: no existing inventory row matched "
                    f"Business Service CI Name '{business_service_ci_name}'.",
                )
                continue

            result.matched_count += 1
            if item.active_users == active_users:
                result.unchanged_count += 1
            else:
                item.active_users = active_users
                result.updated_count += 1

            if result.updated_count and result.updated_count % INGESTION_BATCH_SIZE == 0:
                db.commit()

        recompute_application_ticket_user_metrics(db, project_id)
        db.flush()
        backfill_ticket_hosting_env_from_inventory(db, project_id)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            f"Application inventory active users could not be updated: {exc}"
        ) from exc
    except UnicodeDecodeError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            "Application inventory CSV could not be decoded using supported encodings: "
            + ", ".join(CSV_ENCODING_CANDIDATES)
        ) from exc
    except Exception:
        db.rollback()
        raise

    return result


def update_application_inventory_hosting_env_from_file(
    db: Session,
    project_id: UUID,
    path: Path,
) -> InventoryHostingEnvUpdateResult:
    """Update only hosting_env from an inventory workbook; never insert or upsert rows."""
    ensure_project_exists(db, project_id)
    result = InventoryHostingEnvUpdateResult(project_id=project_id)

    try:
        for parsed_row in iter_inventory_file_rows(path, result):
            result.total_rows += 1
            if not raw_data_has_core_field(parsed_row.raw_data, "hosting_env"):
                raise ApplicationInventoryError("Hosting Env column was not found.")

            business_service_ci_name = text_or_none(
                get_raw_value(parsed_row.raw_data, "business_service_ci_name")
            )
            if business_service_ci_name is None:
                result.skipped_count += 1
                append_sample_message(
                    result.errors,
                    f"Row {parsed_row.row_number}: Business Service CI Name is required.",
                )
                continue

            parent_application_name = text_or_none(
                get_raw_value(parsed_row.raw_data, "parent_application_name")
            )
            assignment_group = text_or_none(get_raw_value(parsed_row.raw_data, "assignment_group"))
            item = db.scalar(
                select(ApplicationInventoryItem)
                .where(
                    inventory_match_condition(
                        project_id,
                        business_service_ci_name=business_service_ci_name,
                        parent_application_name=parent_application_name,
                        assignment_group=assignment_group,
                    )
                )
                .limit(1)
            )
            if item is None:
                fallback_candidates = db.scalars(
                    select(ApplicationInventoryItem)
                    .where(
                        ApplicationInventoryItem.project_id == project_id,
                        ApplicationInventoryItem.is_current.is_(True),
                        func.lower(func.btrim(ApplicationInventoryItem.business_service_ci_name))
                        == normalize_match_key(business_service_ci_name),
                        func.coalesce(
                            func.lower(
                                func.btrim(ApplicationInventoryItem.parent_application_name)
                            ),
                            "",
                        )
                        == (normalize_match_key(parent_application_name) or ""),
                    )
                    .limit(2)
                ).all()
                if len(fallback_candidates) == 1:
                    item = fallback_candidates[0]
                elif len(fallback_candidates) > 1:
                    result.unmatched_count += 1
                    append_sample_message(
                        result.warnings,
                        "Row "
                        f"{parsed_row.row_number}: multiple existing inventory rows matched "
                        f"Business Service CI Name '{business_service_ci_name}' and parent "
                        "application, so Hosting Env was not updated.",
                    )
                    continue
            if item is None:
                result.unmatched_count += 1
                append_sample_message(
                    result.warnings,
                    "Row "
                    f"{parsed_row.row_number}: no existing inventory row matched "
                    f"Business Service CI Name '{business_service_ci_name}'.",
                )
                continue

            hosting_env = text_or_none(get_raw_value(parsed_row.raw_data, "hosting_env"))
            result.matched_count += 1
            if item.hosting_env == hosting_env:
                result.unchanged_count += 1
            else:
                item.hosting_env = hosting_env
                result.updated_count += 1

            if result.updated_count and result.updated_count % INGESTION_BATCH_SIZE == 0:
                db.commit()

        db.flush()
        backfill_ticket_hosting_env_from_inventory(db, project_id)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            f"Application inventory Hosting Env values could not be updated: {exc}"
        ) from exc
    except UnicodeDecodeError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            "Application inventory CSV could not be decoded using supported encodings: "
            + ", ".join(CSV_ENCODING_CANDIDATES)
        ) from exc
    except Exception:
        db.rollback()
        raise

    return result


def iter_inventory_csv_rows(path: Path) -> Iterator[ParsedSourceRow]:
    encoding = detect_csv_encoding(path)
    with path.open("r", encoding=encoding, newline="") as csv_file:
        reader = csv.reader(csv_file)
        try:
            headers = make_unique_headers(next(reader))
        except StopIteration:
            return

        for row_number, values in enumerate(reader, start=2):
            raw_data = build_raw_data(headers, values)
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(row_number=row_number, raw_data=raw_data)


def find_inventory_header_row(rows: Iterator[tuple[Any, ...]]) -> tuple[int, list[str]] | None:
    for row_number, values in enumerate(rows, start=1):
        raw_headers = list(values)
        normalized_headers = {
            normalize_source_column_name(str(value))
            for value in raw_headers
            if value is not None and str(value).strip()
        }
        required_header_aliases = (
            *CORE_FIELD_ALIASES["business_service_ci_name"],
            *CORE_FIELD_ALIASES["assignment_group"],
        )
        if any(
            normalize_source_column_name(alias) in normalized_headers
            for alias in required_header_aliases
        ):
            return row_number, make_unique_headers(raw_headers)
    return None


def iter_inventory_xlsx_rows(
    path: Path,
    result: InventoryUploadResult,
) -> Iterator[ParsedSourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet_name = next(
            (
                sheet_name
                for sheet_name in TARGET_WORKSHEET_NAMES
                if sheet_name in workbook.sheetnames
            ),
            None,
        )
        if worksheet_name is not None:
            worksheet = workbook[worksheet_name]
        else:
            worksheet = workbook.worksheets[0]
            append_sample_message(
                result.warnings,
                "Worksheet 'Group-App-BizService-Inscope' was not found; used first worksheet.",
            )

        rows = worksheet.iter_rows(values_only=True)
        header = find_inventory_header_row(rows)
        if header is None:
            raise ApplicationInventoryError(
                "Could not find a header row containing Support group name or Business "
                "Service CI Name."
            )

        header_row_number, headers = header
        for row_number, values in enumerate(rows, start=header_row_number + 1):
            raw_data = build_raw_data(headers, list(values))
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(
                    row_number=row_number,
                    raw_data=raw_data,
                    source_sheet_name=worksheet.title,
                )
    finally:
        workbook.close()


def iter_inventory_file_rows(
    path: Path,
    result: InventoryUploadResult,
) -> Iterator[ParsedSourceRow]:
    extension = path.suffix.lower()
    if extension == ".csv":
        yield from iter_inventory_csv_rows(path)
        return
    if extension == ".xlsx":
        yield from iter_inventory_xlsx_rows(path, result)
        return
    raise ApplicationInventoryError(
        f"Unsupported application inventory file extension: {extension}. Use CSV or XLSX."
    )


def ensure_project_exists(db: Session, project_id: UUID) -> None:
    if db.get(Project, project_id) is None:
        raise FileNotFoundError(f"Project {project_id} was not found.")


def upload_application_inventory_file(
    db: Session,
    project_id: UUID,
    path: Path,
    source_filename: str,
) -> InventoryUploadResult:
    ensure_project_exists(db, project_id)
    result = InventoryUploadResult(project_id=project_id)

    try:
        replacement_rows: dict[tuple[str, str, str], dict[str, Any]] = {}
        for parsed_row in iter_inventory_file_rows(path, result):
            result.total_rows += 1
            values = clean_inventory_values(
                project_id,
                source_filename,
                parsed_row,
                result,
            )
            if values is None:
                continue

            row_key = inventory_values_key(values)
            if row_key in replacement_rows:
                result.updated_count += 1
            else:
                result.inserted_count += 1
            replacement_rows[row_key] = values

        if not replacement_rows:
            raise ApplicationInventoryError(
                "Application Inventory upload did not contain any valid rows; the active "
                "inventory reference set was not changed."
            )

        replaced_at = datetime.now(UTC)
        db.execute(
            update(ApplicationInventoryItem)
            .where(
                ApplicationInventoryItem.project_id == project_id,
                ApplicationInventoryItem.is_current.is_(True),
            )
            .values(is_current=False, replaced_at=replaced_at)
        )
        for values in replacement_rows.values():
            db.add(ApplicationInventoryItem(**values))

        db.flush()
        sync_application_inventory_scope_from_assignment_groups(db, project_id)
        db.commit()
    except SQLAlchemyError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            f"Application inventory rows could not be saved: {exc}"
        ) from exc
    except UnicodeDecodeError as exc:
        db.rollback()
        raise ApplicationInventoryError(
            "Application inventory CSV could not be decoded using supported encodings: "
            + ", ".join(CSV_ENCODING_CANDIDATES)
        ) from exc
    except Exception:
        db.rollback()
        raise

    return result


def list_inventory_items(db: Session, project_id: UUID) -> list[ApplicationInventoryItem]:
    ensure_project_exists(db, project_id)
    statement = (
        select(ApplicationInventoryItem)
        .where(
            ApplicationInventoryItem.project_id == project_id,
            ApplicationInventoryItem.is_current.is_(True),
        )
        .order_by(
            ApplicationInventoryItem.parent_application_name.asc().nullslast(),
            ApplicationInventoryItem.business_service_ci_name.asc(),
            ApplicationInventoryItem.assignment_group.asc().nullslast(),
        )
    )
    return list(db.scalars(statement).all())


def update_inventory_item(
    db: Session,
    item_id: UUID,
    values: dict[str, Any],
) -> ApplicationInventoryItem:
    item = db.get(ApplicationInventoryItem, item_id)
    if item is None:
        raise FileNotFoundError(f"Application inventory item {item_id} was not found.")

    for field_name, value in values.items():
        if hasattr(item, field_name):
            setattr(item, field_name, value)
    recompute_application_ticket_user_metrics(db, item.project_id)
    db.commit()
    db.refresh(item)
    return item


def deactivate_inventory_item(db: Session, item_id: UUID) -> ApplicationInventoryItem:
    return update_inventory_item(db, item_id, {"active": False})


def reset_inventory_ticket_columns(db: Session, project_id: UUID) -> None:
    db.execute(
        update(Ticket)
        .where(Ticket.project_id == project_id)
        .values(
            application_inventory_id=None,
            parent_application_number=None,
            parent_application_name=None,
            business_service_ci_name=None,
            application_owner=None,
            support_lead=None,
            functional_track=None,
            ams_owner=None,
            supported_by_vendor=None,
            service_type=None,
            service_entitlement=None,
            assignment_group_owner=None,
            sap_non_sap=None,
            architecture_type=None,
            business_critical=None,
            install_type=None,
            hosting_env=None,
            derived_vendor=None,
        )
    )
    db.execute(
        update(AssessmentOutOfScopeTicket)
        .where(AssessmentOutOfScopeTicket.project_id == project_id)
        .values(
            application_inventory_id=None,
            parent_application_number=None,
            parent_application_name=None,
            business_service_ci_name=None,
            application_owner=None,
            support_lead=None,
            functional_track=None,
            ams_owner=None,
            supported_by_vendor=None,
            service_type=None,
            service_entitlement=None,
            assignment_group_owner=None,
            sap_non_sap=None,
            architecture_type=None,
            business_critical=None,
            install_type=None,
            hosting_env=None,
            derived_vendor=None,
        )
    )


def backfill_ticket_hosting_env_from_inventory(db: Session, project_id: UUID) -> int:
    updated_count = 0
    business_critical_expression = business_critical_expression_sql()
    for table_name in ("tickets", "assessment_out_of_scope_tickets"):
        result = db.execute(
            text(
                f"""
                UPDATE {table_name} AS t
                SET
                    hosting_env = NULLIF(btrim(i.hosting_env), ''),
                    business_critical = {business_critical_expression}
                FROM application_inventory_items AS i
                WHERE t.project_id = CAST(:project_id AS uuid)
                  AND t.application_inventory_id = i.id
                  AND i.is_current IS true
                  AND (
                    t.hosting_env IS DISTINCT FROM NULLIF(btrim(i.hosting_env), '')
                    OR t.business_critical IS DISTINCT FROM {business_critical_expression}
                  )
                """,
            ),
            {"project_id": str(project_id)},
        )
        updated_count += int(result.rowcount or 0)
    return updated_count


def business_critical_expression_sql(alias: str = "i") -> str:
    keys = ("Business criticality", "Biz Criticality", "Business Criticality", "Business Critical")
    expressions = []
    for key in keys:
        escaped_key = key.replace("'", "''")
        expressions.append(f"NULLIF(btrim({alias}.cmdb_payload ->> '{escaped_key}'), '')")
    return f"COALESCE({', '.join(expressions)})"


def cmdb_payload_text_expression_sql(alias: str, *keys: str) -> str:
    expressions = []
    for key in keys:
        escaped_key = key.replace("'", "''")
        expressions.append(f"NULLIF(btrim({alias}.cmdb_payload ->> '{escaped_key}'), '')")
    return f"COALESCE({', '.join(expressions)})"


def normalized_business_service_ci_sql(expression: str) -> str:
    return (
        "lower(regexp_replace("
        f"btrim(replace({expression}, chr(160), ' ')), "
        "'[[:space:]]+', ' ', 'g'"
        "))"
    )


def normalized_text_key_sql(expression: str) -> str:
    return normalized_business_service_ci_sql(expression)


def sync_application_inventory_scope_from_assignment_groups(
    db: Session,
    project_id: UUID,
) -> int:
    inventory_assignment_group_key = normalized_text_key_sql("i.assignment_group")
    scope_assignment_group_key = normalized_text_key_sql("s.assignment_group")
    db.execute(
        text(
            """
            UPDATE application_inventory_items
            SET scope_status = 'unknown'
            WHERE project_id = CAST(:project_id AS uuid)
              AND is_current IS true
            """
        ),
        {"project_id": str(project_id)},
    )
    result = db.execute(
        text(
            f"""
            WITH scope_rows AS (
                SELECT
                    {scope_assignment_group_key} AS assignment_group_key,
                    functional_track,
                    is_in_scope
                FROM in_scope_assignment_groups AS s
                WHERE s.project_id = CAST(:project_id AS uuid)
                  AND s.is_active IS true
                  AND NULLIF(btrim(s.assignment_group), '') IS NOT NULL
            )
            UPDATE application_inventory_items AS i
            SET
                scope_status = CASE
                    WHEN scope_rows.is_in_scope IS true THEN 'in_scope'
                    ELSE 'out_of_scope'
                END,
                functional_track = scope_rows.functional_track
            FROM scope_rows
            WHERE i.project_id = CAST(:project_id AS uuid)
              AND i.is_current IS true
              AND NULLIF(btrim(i.assignment_group), '') IS NOT NULL
              AND {inventory_assignment_group_key} = scope_rows.assignment_group_key
            """
        ),
        {"project_id": str(project_id)},
    )
    return int(result.rowcount or 0)


def update_tickets_from_inventory(
    db: Session,
    project_id: UUID,
    *,
    ticket_column: str,
    table_name: str = "tickets",
    ticket_type: str | None = None,
) -> int:
    if table_name not in {"tickets", "assessment_out_of_scope_tickets"}:
        raise ApplicationInventoryError(f"Unsupported enrichment table: {table_name}")
    ticket_key_expression = normalized_business_service_ci_sql(f"t.{ticket_column}")
    inventory_key_expression = normalized_business_service_ci_sql("i.business_service_ci_name")
    ticket_type_filter = ""
    parameters = {"project_id": str(project_id)}
    if ticket_type is not None:
        ticket_type_filter = "AND t.ticket_type = :ticket_type"
        parameters["ticket_type"] = ticket_type
    architecture_expression = cmdb_payload_text_expression_sql(
        "i",
        "Architecture type",
        "Architecture Type",
    )
    install_type_expression = cmdb_payload_text_expression_sql(
        "i",
        "Install type",
        "Install Type",
        "Install status",
        "Install Status",
    )

    statement = text(
        f"""
        WITH candidates AS (
            SELECT
                t.id AS ticket_id,
                i.id AS inventory_id,
                i.application_number_apm,
                i.parent_application_name,
                i.business_service_ci_name,
                i.application_owner,
                i.support_lead,
                i.supported_by_vendor,
                i.service_type,
                i.service_entitlement,
                i.assignment_group_owner,
                i.sap_non_sap,
                {architecture_expression} AS architecture_type,
                {install_type_expression} AS install_type,
                i.hosting_env,
                {business_critical_expression_sql()} AS business_critical,
                row_number() OVER (
                    PARTITION BY t.id
                    ORDER BY
                        CASE
                            WHEN i.active IS true THEN 0
                            WHEN i.active IS NULL THEN 1
                            ELSE 2
                        END,
                        CASE
                            WHEN lower(coalesce(btrim(t.assignment_group), '')) =
                                 lower(coalesce(btrim(i.assignment_group), ''))
                                 AND nullif(btrim(i.assignment_group), '') IS NOT NULL
                            THEN 0
                            ELSE 1
                        END,
                        i.source_row_number ASC NULLS LAST,
                        i.created_at ASC,
                        i.id ASC
                ) AS row_rank
            FROM {table_name} AS t
            JOIN application_inventory_items AS i
              ON i.project_id = t.project_id
             AND i.is_current IS true
             AND i.active IS NOT false
             AND nullif(btrim(t.{ticket_column}), '') IS NOT NULL
             AND {ticket_key_expression} = {inventory_key_expression}
            WHERE t.project_id = CAST(:project_id AS uuid)
              {ticket_type_filter}
              AND t.application_inventory_id IS NULL
        )
        UPDATE {table_name} AS t
        SET
            application_inventory_id = candidates.inventory_id,
            parent_application_number = candidates.application_number_apm,
            parent_application_name = candidates.parent_application_name,
            business_service_ci_name = candidates.business_service_ci_name,
            application_owner = candidates.application_owner,
            support_lead = candidates.support_lead,
            supported_by_vendor = candidates.supported_by_vendor,
            service_type = candidates.service_type,
            service_entitlement = candidates.service_entitlement,
            assignment_group_owner = candidates.assignment_group_owner,
            business_critical = candidates.business_critical,
            hosting_env = candidates.hosting_env,
            sap_non_sap = candidates.sap_non_sap,
            architecture_type = candidates.architecture_type,
            install_type = candidates.install_type,
            derived_vendor = candidates.supported_by_vendor
        FROM candidates
        WHERE candidates.row_rank = 1
          AND t.id = candidates.ticket_id
        """
    )
    result = db.execute(statement, parameters)
    return int(result.rowcount or 0)


def update_ticket_support_group_fields_from_inventory(
    db: Session,
    project_id: UUID,
    *,
    table_name: str = "tickets",
    only_unmatched: bool = False,
) -> int:
    if table_name not in {"tickets", "assessment_out_of_scope_tickets"}:
        raise ApplicationInventoryError(f"Unsupported enrichment table: {table_name}")
    ticket_assignment_group_key = normalized_text_key_sql("t.assignment_group")
    scope_assignment_group_key = normalized_text_key_sql("assignment_group")
    unmatched_filter = "AND t.application_inventory_id IS NULL" if only_unmatched else ""
    statement = text(
        f"""
        WITH scope_values AS (
            SELECT
                {scope_assignment_group_key} AS assignment_group_key,
                functional_track
            FROM in_scope_assignment_groups
            WHERE project_id = CAST(:project_id AS uuid)
              AND is_active IS true
              AND NULLIF(btrim(assignment_group), '') IS NOT NULL
        )
        UPDATE {table_name} AS t
        SET
            functional_track = scope_values.functional_track
        FROM scope_values
        WHERE t.project_id = CAST(:project_id AS uuid)
          AND NULLIF(btrim(t.assignment_group), '') IS NOT NULL
          {unmatched_filter}
          AND {ticket_assignment_group_key} = scope_values.assignment_group_key
          AND t.functional_track IS DISTINCT FROM scope_values.functional_track
        """
    )
    result = db.execute(statement, {"project_id": str(project_id)})
    return int(result.rowcount or 0)


def count_tickets(db: Session, project_id: UUID, *, matched: bool | None = None) -> int:
    statement = select(func.count(Ticket.id)).where(Ticket.project_id == project_id)
    if matched is True:
        statement = statement.where(Ticket.application_inventory_id.is_not(None))
    elif matched is False:
        statement = statement.where(Ticket.application_inventory_id.is_(None))
    return int(db.scalar(statement) or 0)


def distinct_ticket_business_service_count(db: Session, project_id: UUID) -> int:
    statement = select(
        func.count(func.distinct(func.lower(func.btrim(Ticket.business_service))))
    ).where(
        Ticket.project_id == project_id,
        Ticket.business_service.is_not(None),
        func.btrim(Ticket.business_service) != "",
    )
    return int(db.scalar(statement) or 0)


def distinct_inventory_business_service_count(db: Session, project_id: UUID) -> int:
    statement = select(
        func.count(
            func.distinct(func.lower(func.btrim(ApplicationInventoryItem.business_service_ci_name)))
        )
    ).where(
        ApplicationInventoryItem.project_id == project_id,
        ApplicationInventoryItem.is_current.is_(True),
        ApplicationInventoryItem.business_service_ci_name.is_not(None),
        func.btrim(ApplicationInventoryItem.business_service_ci_name) != "",
    )
    return int(db.scalar(statement) or 0)


def matched_ticket_business_service_count(db: Session, project_id: UUID) -> int:
    statement = text(
        """
        WITH ticket_services AS (
            SELECT DISTINCT lower(btrim(business_service)) AS service_key
            FROM tickets
            WHERE project_id = CAST(:project_id AS uuid)
              AND nullif(btrim(business_service), '') IS NOT NULL
        ),
        inventory_services AS (
            SELECT DISTINCT lower(btrim(business_service_ci_name)) AS service_key
            FROM application_inventory_items
            WHERE project_id = CAST(:project_id AS uuid)
              AND is_current IS true
              AND nullif(btrim(business_service_ci_name), '') IS NOT NULL
        )
        SELECT count(*) AS matched_count
        FROM ticket_services t
        JOIN inventory_services i ON i.service_key = t.service_key
        """
    )
    return int(db.execute(statement, {"project_id": str(project_id)}).scalar_one() or 0)


def top_unmatched_values(db: Session, project_id: UUID, column: Any) -> list[ValueCount]:
    statement = (
        select(column, func.count(Ticket.id))
        .where(
            Ticket.project_id == project_id,
            Ticket.application_inventory_id.is_(None),
            column.is_not(None),
            func.btrim(column) != "",
        )
        .group_by(column)
        .order_by(func.count(Ticket.id).desc(), column.asc())
        .limit(TOP_UNMATCHED_LIMIT)
    )
    return [
        ValueCount(value=str(value), count=int(count))
        for value, count in db.execute(statement).all()
        if value
    ]


def current_business_service_match_count(db: Session, project_id: UUID) -> int:
    statement = select(func.count(Ticket.id)).where(
        Ticket.project_id == project_id,
        Ticket.application_inventory_id.is_not(None),
        Ticket.business_service_ci_name.is_not(None),
        func.lower(func.btrim(Ticket.business_service))
        == func.lower(func.btrim(Ticket.business_service_ci_name)),
    )
    return int(db.scalar(statement) or 0)


def current_application_match_count(db: Session, project_id: UUID) -> int:
    statement = select(func.count(Ticket.id)).where(
        Ticket.project_id == project_id,
        Ticket.application_inventory_id.is_not(None),
        Ticket.business_service_ci_name.is_not(None),
        func.lower(func.btrim(Ticket.application))
        == func.lower(func.btrim(Ticket.business_service_ci_name)),
        func.coalesce(func.lower(func.btrim(Ticket.business_service)), "")
        != func.lower(func.btrim(Ticket.business_service_ci_name)),
    )
    return int(db.scalar(statement) or 0)


def calculate_rate(numerator: int, denominator: int) -> float | None:
    if denominator == 0:
        return None
    return round((numerator / denominator) * 100, 2)


def build_inventory_enrichment_summary(
    db: Session,
    project_id: UUID,
    *,
    updated_tickets: int = 0,
    matched_by_business_service_count: int | None = None,
    matched_by_application_count: int | None = None,
) -> InventoryEnrichmentSummary:
    ensure_project_exists(db, project_id)
    total_tickets = count_tickets(db, project_id)
    matched_tickets = count_tickets(db, project_id, matched=True)
    unmatched_tickets = max(total_tickets - matched_tickets, 0)
    distinct_ticket_services = distinct_ticket_business_service_count(db, project_id)
    distinct_inventory_services = distinct_inventory_business_service_count(db, project_id)
    matched_services = matched_ticket_business_service_count(db, project_id)
    unmatched_services = max(distinct_ticket_services - matched_services, 0)

    return InventoryEnrichmentSummary(
        project_id=project_id,
        total_tickets=total_tickets,
        matched_tickets=matched_tickets,
        unmatched_tickets=unmatched_tickets,
        updated_tickets=updated_tickets,
        match_rate_pct=calculate_rate(matched_tickets, total_tickets),
        matched_by_business_service_count=(
            current_business_service_match_count(db, project_id)
            if matched_by_business_service_count is None
            else matched_by_business_service_count
        ),
        matched_by_application_count=(
            current_application_match_count(db, project_id)
            if matched_by_application_count is None
            else matched_by_application_count
        ),
        unmatched_business_service_count=unmatched_services,
        distinct_ticket_business_service_count=distinct_ticket_services,
        distinct_inventory_business_service_count=distinct_inventory_services,
        top_unmatched_business_services=top_unmatched_values(
            db,
            project_id,
            Ticket.business_service,
        ),
        top_unmatched_applications=top_unmatched_values(db, project_id, Ticket.application),
        top_unmatched_assignment_groups=top_unmatched_values(
            db,
            project_id,
            Ticket.assignment_group,
        ),
    )


def enrich_tickets_from_inventory(
    db: Session,
    project_id: UUID,
    *,
    replace_existing: bool,
) -> InventoryEnrichmentSummary:
    ensure_project_exists(db, project_id)
    if replace_existing:
        reset_inventory_ticket_columns(db, project_id)
        db.flush()

    update_ticket_support_group_fields_from_inventory(
        db,
        project_id,
        only_unmatched=True,
    )
    update_ticket_support_group_fields_from_inventory(
        db,
        project_id,
        table_name="assessment_out_of_scope_tickets",
        only_unmatched=True,
    )
    business_service_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="business_service",
        ticket_type="INCIDENT",
    )
    sc_task_cmdb_ci_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="cmdb_ci",
        ticket_type="SERVICE_CATALOG_TASK",
    )
    application_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="application",
        ticket_type="INCIDENT",
    )
    out_business_service_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="business_service",
        table_name="assessment_out_of_scope_tickets",
        ticket_type="INCIDENT",
    )
    out_sc_task_cmdb_ci_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="cmdb_ci",
        table_name="assessment_out_of_scope_tickets",
        ticket_type="SERVICE_CATALOG_TASK",
    )
    out_application_updates = update_tickets_from_inventory(
        db,
        project_id,
        ticket_column="application",
        table_name="assessment_out_of_scope_tickets",
        ticket_type="INCIDENT",
    )
    recompute_application_ticket_user_metrics(db, project_id)
    db.commit()

    return build_inventory_enrichment_summary(
        db,
        project_id,
        updated_tickets=(
            business_service_updates
            + sc_task_cmdb_ci_updates
            + application_updates
            + out_business_service_updates
            + out_sc_task_cmdb_ci_updates
            + out_application_updates
        ),
        matched_by_business_service_count=business_service_updates,
        matched_by_application_count=application_updates,
    )


def inventory_filter_values(db: Session, project_id: UUID) -> dict[str, list[str]]:
    ensure_project_exists(db, project_id)
    columns = {
        "application_owners": ApplicationInventoryItem.application_owner,
        "support_leads": ApplicationInventoryItem.support_lead,
        "functional_tracks": ApplicationInventoryItem.functional_track,
        "ams_owners": ApplicationInventoryItem.ams_owner,
        "supported_by_vendors": ApplicationInventoryItem.supported_by_vendor,
        "hosting_envs": ApplicationInventoryItem.hosting_env,
        "parent_application_names": ApplicationInventoryItem.parent_application_name,
        "business_service_ci_names": ApplicationInventoryItem.business_service_ci_name,
        "assignment_groups": ApplicationInventoryItem.assignment_group,
    }
    values: dict[str, list[str]] = {}
    for key, column in columns.items():
        statement = (
            select(column)
            .distinct()
            .where(
                ApplicationInventoryItem.project_id == project_id,
                ApplicationInventoryItem.is_current.is_(True),
                column.is_not(None),
                func.btrim(column) != "",
            )
            .order_by(column)
        )
        values[key] = [str(value) for value in db.scalars(statement).all() if value]
    return values


def unmatched_business_services(
    db: Session,
    project_id: UUID,
    *,
    limit: int,
    offset: int,
) -> BusinessServiceCoverage:
    ensure_project_exists(db, project_id)
    distinct_ticket_services = distinct_ticket_business_service_count(db, project_id)
    distinct_inventory_services = distinct_inventory_business_service_count(db, project_id)
    matched_services = matched_ticket_business_service_count(db, project_id)
    unmatched_services = max(distinct_ticket_services - matched_services, 0)

    statement = text(
        """
        WITH inventory_services AS (
            SELECT DISTINCT lower(btrim(business_service_ci_name)) AS service_key
            FROM application_inventory_items
            WHERE project_id = CAST(:project_id AS uuid)
              AND is_current IS true
              AND nullif(btrim(business_service_ci_name), '') IS NOT NULL
        )
        SELECT
            t.business_service,
            count(*) AS ticket_count,
            count(DISTINCT t.assignment_group) AS assignment_group_count,
            (array_remove(array_agg(DISTINCT t.assignment_group), NULL))[1:5]
                AS sample_assignment_groups,
            (array_agg(t.ticket_number ORDER BY t.ticket_number))[1:5]
                AS sample_ticket_numbers
        FROM tickets AS t
        LEFT JOIN inventory_services AS i
          ON i.service_key = lower(btrim(t.business_service))
        WHERE t.project_id = CAST(:project_id AS uuid)
          AND nullif(btrim(t.business_service), '') IS NOT NULL
          AND i.service_key IS NULL
        GROUP BY t.business_service
        ORDER BY count(*) DESC, t.business_service ASC
        LIMIT :limit OFFSET :offset
        """
    )
    rows = [
        UnmatchedBusinessService(
            business_service=str(row["business_service"]),
            ticket_count=int(row["ticket_count"] or 0),
            assignment_group_count=int(row["assignment_group_count"] or 0),
            sample_assignment_groups=list(row["sample_assignment_groups"] or []),
            sample_ticket_numbers=list(row["sample_ticket_numbers"] or []),
        )
        for row in db.execute(
            statement,
            {"project_id": str(project_id), "limit": limit, "offset": offset},
        )
        .mappings()
        .all()
    ]

    return BusinessServiceCoverage(
        project_id=project_id,
        distinct_ticket_business_service_count=distinct_ticket_services,
        distinct_inventory_business_service_count=distinct_inventory_services,
        matched_business_service_count=matched_services,
        unmatched_business_service_count=unmatched_services,
        business_service_coverage_pct=calculate_rate(matched_services, distinct_ticket_services),
        rows=rows,
    )


def count_distinct_nonblank(db: Session, project_id: UUID, model: Any, column: Any) -> int:
    statement = select(func.count(func.distinct(func.lower(func.btrim(column))))).where(
        model.project_id == project_id,
        column.is_not(None),
        func.btrim(column) != "",
    )
    return int(db.scalar(statement) or 0)


def count_distinct_ticket_scope_nonblank(
    db: Session,
    project_id: UUID,
    column: Any,
    *,
    is_in_scope: bool,
) -> int:
    statement = select(func.count(func.distinct(func.lower(func.btrim(column))))).where(
        Ticket.project_id == project_id,
        Ticket.is_in_scope.is_(is_in_scope),
        column.is_not(None),
        func.btrim(column) != "",
    )
    return int(db.scalar(statement) or 0)


def top_out_of_scope_values(db: Session, project_id: UUID, column: Any) -> list[ValueCount]:
    statement = (
        select(column, func.count(Ticket.id))
        .where(
            Ticket.project_id == project_id,
            Ticket.is_in_scope.is_(False),
            column.is_not(None),
            func.btrim(column) != "",
        )
        .group_by(column)
        .order_by(func.count(Ticket.id).desc(), column.asc())
        .limit(TOP_UNMATCHED_LIMIT)
    )
    return [
        ValueCount(value=str(value), count=int(count))
        for value, count in db.execute(statement).all()
        if value
    ]


def build_scope_summary(db: Session, project_id: UUID) -> ScopeSummary:
    ensure_project_exists(db, project_id)
    in_scope_tickets = int(
        db.scalar(
            select(func.count(Ticket.id)).where(
                Ticket.project_id == project_id,
                Ticket.is_in_scope.is_(True),
            )
        )
        or 0
    )
    out_of_scope_tickets = int(
        db.scalar(
            select(func.count(Ticket.id)).where(
                Ticket.project_id == project_id,
                Ticket.is_in_scope.is_(False),
            )
        )
        or 0
    )
    total_classified_tickets = in_scope_tickets + out_of_scope_tickets

    return ScopeSummary(
        project_id=project_id,
        in_scope_tickets=in_scope_tickets,
        out_of_scope_tickets=out_of_scope_tickets,
        total_classified_tickets=total_classified_tickets,
        in_scope_pct=calculate_rate(in_scope_tickets, total_classified_tickets),
        out_of_scope_pct=calculate_rate(out_of_scope_tickets, total_classified_tickets),
        distinct_in_scope_assignment_groups=count_distinct_ticket_scope_nonblank(
            db,
            project_id,
            Ticket.assignment_group,
            is_in_scope=True,
        ),
        distinct_out_of_scope_assignment_groups=count_distinct_ticket_scope_nonblank(
            db,
            project_id,
            Ticket.assignment_group,
            is_in_scope=False,
        ),
        top_out_of_scope_assignment_groups=top_out_of_scope_values(
            db,
            project_id,
            Ticket.assignment_group,
        ),
        top_out_of_scope_business_services=top_out_of_scope_values(
            db,
            project_id,
            Ticket.business_service,
        ),
    )
