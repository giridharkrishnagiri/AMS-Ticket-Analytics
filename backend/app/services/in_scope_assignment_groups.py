from __future__ import annotations

import csv
import re
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select, text
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models import InScopeAssignmentGroup, Project
from app.services.ingestion import (
    CSV_ENCODING_CANDIDATES,
    ParsedSourceRow,
    build_raw_data,
    detect_csv_encoding,
    make_unique_headers,
    normalize_source_column_name,
    row_has_any_value,
)

MAX_MESSAGE_SAMPLES = 50

ASSIGNMENT_GROUP_ALIASES = (
    "Assigment Groups",
    "Assignment Groups",
    "Assignment Group",
    "Assignment group",
    "assignment_group",
    "Support Group",
    "Support group",
    "Group",
)
FUNCTIONAL_TRACK_ALIASES = (
    "Track",
    "Functional Track",
    "functional_track",
    "Functional track",
)
IN_SCOPE_ALIASES = (
    "In scope",
    "In Scope",
    "In-Scope",
    "InScope",
    "is_in_scope",
    "Scope",
)


class InScopeAssignmentGroupsError(Exception):
    pass


def text_or_none(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.replace("\xa0", " ").strip()
        return text or None
    text = str(value).replace("\xa0", " ").strip()
    return text or None


@dataclass(frozen=True)
class InScopeAssignmentGroupPreviewRow:
    assignment_group: str
    functional_track: str | None
    is_in_scope: bool
    source_row_number: int | None


@dataclass
class InScopeAssignmentGroupsImportResult:
    project_id: UUID
    source_filename: str
    total_rows: int = 0
    imported_count: int = 0
    skipped_count: int = 0
    duplicate_count: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    preview_rows: list[InScopeAssignmentGroupPreviewRow] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass(frozen=True)
class InScopeAssignmentGroupsStatus:
    project_id: UUID
    active_count: int
    last_imported_at: Any | None
    preview_rows: list[InScopeAssignmentGroupPreviewRow]


@dataclass(frozen=True)
class AssignmentGroupScope:
    assignment_group: str
    functional_track: str | None
    is_in_scope: bool


@dataclass(frozen=True)
class AssignmentGroupScopeUpdate:
    id: UUID
    functional_track: str | None
    is_in_scope: bool


@dataclass(frozen=True)
class AssignmentGroupScopeChange:
    id: UUID
    assignment_group: str
    previous_functional_track: str | None
    next_functional_track: str | None
    previous_is_in_scope: bool
    next_is_in_scope: bool
    tickets_updated: int


@dataclass(frozen=True)
class AssignmentGroupScopeUpdateResult:
    project_id: UUID
    submitted_count: int
    changed_count: int
    unchanged_count: int
    tickets_updated_count: int
    missing_count: int
    changes: list[AssignmentGroupScopeChange]
    warnings: list[str]


def append_sample_message(messages: list[str], message: str) -> None:
    if len(messages) < MAX_MESSAGE_SAMPLES:
        messages.append(message)


def normalize_assignment_group_key(value: Any) -> str | None:
    text = text_or_none(value)
    if text is None:
        return None
    normalized = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip().casefold()
    return normalized or None


def parse_in_scope_value(value: Any) -> bool | None:
    text = text_or_none(value)
    if text is None:
        return None
    normalized = " ".join(text.replace("-", " ").split()).casefold()
    if normalized in {"yes", "y", "true", "t", "1", "in scope", "inscope"}:
        return True
    if normalized in {
        "no",
        "n",
        "false",
        "f",
        "0",
        "out of scope",
        "out scope",
        "outofscope",
    }:
        return False
    return None


def header_lookup(raw_data: dict[str, Any]) -> dict[str, Any]:
    return {
        normalize_source_column_name(column_name): value
        for column_name, value in raw_data.items()
    }


def get_aliased_value(raw_data: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        if alias in raw_data:
            return raw_data[alias]
    lookup = header_lookup(raw_data)
    for alias in aliases:
        normalized_alias = normalize_source_column_name(alias)
        if normalized_alias in lookup:
            return lookup[normalized_alias]
    return None


def has_alias(headers: set[str], aliases: tuple[str, ...]) -> bool:
    return any(normalize_source_column_name(alias) in headers for alias in aliases)


def validate_headers(raw_data: dict[str, Any]) -> None:
    headers = {normalize_source_column_name(column_name) for column_name in raw_data}
    missing: list[str] = []
    if not has_alias(headers, ASSIGNMENT_GROUP_ALIASES):
        missing.append("Assigment Groups / Assignment Group")
    if not has_alias(headers, FUNCTIONAL_TRACK_ALIASES):
        missing.append("Track / Functional Track")
    if missing:
        raise InScopeAssignmentGroupsError(
            "In-scope assignment group workbook is missing required column(s): "
            + ", ".join(missing)
            + ".",
        )


def iter_reference_csv_rows(path: Path) -> Iterator[ParsedSourceRow]:
    encoding = detect_csv_encoding(path)
    with path.open("r", encoding=encoding, newline="") as csv_file:
        reader = csv.reader(csv_file)
        try:
            headers = make_unique_headers(next(reader))
        except StopIteration:
            return

        for row_number, values in enumerate(reader, start=2):
            raw_data = build_raw_data(headers, values)
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(row_number=row_number, raw_data=raw_data)


def iter_reference_xlsx_rows(path: Path) -> Iterator[ParsedSourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = make_unique_headers(next(rows))
        except StopIteration:
            return
        for row_number, values in enumerate(rows, start=2):
            raw_data = build_raw_data(headers, list(values))
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(row_number=row_number, raw_data=raw_data)
    finally:
        workbook.close()


def iter_reference_file_rows(path: Path) -> Iterator[ParsedSourceRow]:
    extension = path.suffix.lower()
    if extension == ".csv":
        yield from iter_reference_csv_rows(path)
        return
    if extension == ".xlsx":
        yield from iter_reference_xlsx_rows(path)
        return
    raise InScopeAssignmentGroupsError(
        f"Unsupported in-scope assignment group reference extension: {extension}. Use XLSX or CSV.",
    )


def ensure_project(db: Session, project_id: UUID) -> Project:
    project = db.get(Project, project_id)
    if project is None:
        raise FileNotFoundError(f"Project {project_id} was not found.")
    return project


def import_in_scope_assignment_groups(
    db: Session,
    project_id: UUID,
    path: Path,
    source_filename: str,
) -> InScopeAssignmentGroupsImportResult:
    project = ensure_project(db, project_id)
    result = InScopeAssignmentGroupsImportResult(
        project_id=project_id,
        source_filename=source_filename,
    )
    rows_by_key: dict[str, dict[str, Any]] = {}

    try:
        first_row_checked = False
        in_scope_column_present = False
        for parsed_row in iter_reference_file_rows(path):
            result.total_rows += 1
            if not first_row_checked:
                validate_headers(parsed_row.raw_data)
                headers = {
                    normalize_source_column_name(column_name)
                    for column_name in parsed_row.raw_data
                }
                in_scope_column_present = has_alias(headers, IN_SCOPE_ALIASES)
                first_row_checked = True

            assignment_group = text_or_none(
                get_aliased_value(parsed_row.raw_data, ASSIGNMENT_GROUP_ALIASES)
            )
            if assignment_group is None:
                result.skipped_count += 1
                continue

            assignment_group_key = normalize_assignment_group_key(assignment_group)
            if assignment_group_key is None:
                result.skipped_count += 1
                continue

            functional_track = text_or_none(
                get_aliased_value(parsed_row.raw_data, FUNCTIONAL_TRACK_ALIASES)
            )
            if functional_track is None:
                append_sample_message(
                    result.warnings,
                    f"Row {parsed_row.row_number}: Track is blank for '{assignment_group}'.",
                )
            raw_in_scope = get_aliased_value(parsed_row.raw_data, IN_SCOPE_ALIASES)
            parsed_in_scope = parse_in_scope_value(raw_in_scope)
            if in_scope_column_present and parsed_in_scope is None:
                append_sample_message(
                    result.warnings,
                    f"Row {parsed_row.row_number}: In scope is blank or invalid for "
                    f"'{assignment_group}'; defaulted to No.",
                )
            is_in_scope = parsed_in_scope if in_scope_column_present else True
            if is_in_scope is None:
                is_in_scope = False

            if assignment_group_key in rows_by_key:
                result.duplicate_count += 1
            rows_by_key[assignment_group_key] = {
                "client_id": project.client_id,
                "project_id": project_id,
                "assignment_group": assignment_group,
                "assignment_group_key": assignment_group_key,
                "functional_track": functional_track,
                "is_in_scope": is_in_scope,
                "source_filename": source_filename,
                "source_row_number": parsed_row.row_number,
                "is_active": True,
            }

        if result.total_rows == 0:
            raise InScopeAssignmentGroupsError(
                "In-scope assignment group reference file did not contain any data rows.",
            )

        db.execute(
            delete(InScopeAssignmentGroup).where(
                InScopeAssignmentGroup.project_id == project_id,
            )
        )
        for values in rows_by_key.values():
            db.add(InScopeAssignmentGroup(**values))

        result.imported_count = len(rows_by_key)
        result.preview_rows = [
            InScopeAssignmentGroupPreviewRow(
                assignment_group=str(values["assignment_group"]),
                functional_track=values["functional_track"],
                is_in_scope=bool(values["is_in_scope"]),
                source_row_number=values["source_row_number"],
            )
            for values in list(rows_by_key.values())[:25]
        ]
        db.flush()
    except SQLAlchemyError as exc:
        db.rollback()
        raise InScopeAssignmentGroupsError(
            f"In-scope assignment group reference could not be saved: {exc}",
        ) from exc
    except UnicodeDecodeError as exc:
        db.rollback()
        raise InScopeAssignmentGroupsError(
            "In-scope assignment group reference CSV could not be decoded using supported "
            f"encodings: {', '.join(CSV_ENCODING_CANDIDATES)}",
        ) from exc
    except Exception:
        db.rollback()
        raise

    return result


def active_assignment_group_keys(db: Session, project_id: UUID) -> set[str]:
    ensure_project(db, project_id)
    return {
        key
        for key in db.scalars(
            select(InScopeAssignmentGroup.assignment_group_key).where(
                InScopeAssignmentGroup.project_id == project_id,
                InScopeAssignmentGroup.is_active.is_(True),
                InScopeAssignmentGroup.is_in_scope.is_(True),
            ),
        ).all()
        if key
    }


def assignment_group_scope_map(
    db: Session,
    project_id: UUID,
) -> dict[str, AssignmentGroupScope]:
    ensure_project(db, project_id)
    rows = db.scalars(
        select(InScopeAssignmentGroup).where(
            InScopeAssignmentGroup.project_id == project_id,
            InScopeAssignmentGroup.is_active.is_(True),
        )
    ).all()
    return {
        row.assignment_group_key: AssignmentGroupScope(
            assignment_group=row.assignment_group,
            functional_track=row.functional_track,
            is_in_scope=row.is_in_scope,
        )
        for row in rows
        if row.assignment_group_key
    }


def in_scope_assignment_groups_status(
    db: Session,
    project_id: UUID,
    *,
    preview_limit: int = 25,
) -> InScopeAssignmentGroupsStatus:
    ensure_project(db, project_id)
    active_count = int(
        db.scalar(
            select(func.count(InScopeAssignmentGroup.id)).where(
                InScopeAssignmentGroup.project_id == project_id,
                InScopeAssignmentGroup.is_active.is_(True),
            )
        )
        or 0
    )
    last_imported_at = db.scalar(
        select(func.max(InScopeAssignmentGroup.created_at)).where(
            InScopeAssignmentGroup.project_id == project_id,
            InScopeAssignmentGroup.is_active.is_(True),
        )
    )
    preview = db.scalars(
        select(InScopeAssignmentGroup)
        .where(
            InScopeAssignmentGroup.project_id == project_id,
            InScopeAssignmentGroup.is_active.is_(True),
        )
        .order_by(
            InScopeAssignmentGroup.assignment_group.asc(),
            InScopeAssignmentGroup.id.asc(),
        )
        .limit(preview_limit),
    ).all()
    return InScopeAssignmentGroupsStatus(
        project_id=project_id,
        active_count=active_count,
        last_imported_at=last_imported_at,
        preview_rows=[
            InScopeAssignmentGroupPreviewRow(
                assignment_group=row.assignment_group,
                functional_track=row.functional_track,
                is_in_scope=row.is_in_scope,
                source_row_number=row.source_row_number,
            )
            for row in preview
        ],
    )


def list_in_scope_assignment_groups(
    db: Session,
    project_id: UUID,
    *,
    limit: int = 100,
    offset: int = 0,
) -> list[InScopeAssignmentGroup]:
    ensure_project(db, project_id)
    return list(
        db.scalars(
            select(InScopeAssignmentGroup)
            .where(
                InScopeAssignmentGroup.project_id == project_id,
                InScopeAssignmentGroup.is_active.is_(True),
            )
            .order_by(
                InScopeAssignmentGroup.assignment_group.asc(),
                InScopeAssignmentGroup.id.asc(),
            )
            .offset(offset)
            .limit(limit),
        ).all()
    )


def update_tickets_for_scope_change(
    db: Session,
    *,
    project_id: UUID,
    assignment_group_key: str,
    functional_track: str | None,
    is_in_scope: bool,
) -> int:
    result = db.execute(
        text(
            """
            UPDATE tickets
            SET
                is_in_scope = :is_in_scope,
                functional_track = :functional_track,
                record_updated_at = now()
            WHERE project_id = CAST(:project_id AS uuid)
              AND NULLIF(btrim(assignment_group), '') IS NOT NULL
              AND lower(regexp_replace(btrim(assignment_group), '\\s+', ' ', 'g'))
                  = :assignment_group_key
            """
        ),
        {
            "project_id": str(project_id),
            "assignment_group_key": assignment_group_key,
            "functional_track": functional_track,
            "is_in_scope": is_in_scope,
        },
    )
    return int(result.rowcount or 0)


def update_assignment_group_scope_rows(
    db: Session,
    project_id: UUID,
    updates: list[AssignmentGroupScopeUpdate],
) -> AssignmentGroupScopeUpdateResult:
    ensure_project(db, project_id)
    submitted_count = len(updates)
    if submitted_count == 0:
        return AssignmentGroupScopeUpdateResult(
            project_id=project_id,
            submitted_count=0,
            changed_count=0,
            unchanged_count=0,
            tickets_updated_count=0,
            missing_count=0,
            changes=[],
            warnings=[],
        )

    update_by_id = {update.id: update for update in updates}
    rows = db.scalars(
        select(InScopeAssignmentGroup).where(
            InScopeAssignmentGroup.project_id == project_id,
            InScopeAssignmentGroup.id.in_(list(update_by_id)),
            InScopeAssignmentGroup.is_active.is_(True),
        )
    ).all()
    rows_by_id = {row.id: row for row in rows}
    missing_ids = [row_id for row_id in update_by_id if row_id not in rows_by_id]
    warnings: list[str] = []
    if missing_ids:
        append_sample_message(
            warnings,
            f"{len(missing_ids)} submitted assignment group scope row(s) were not found.",
        )

    changes: list[AssignmentGroupScopeChange] = []
    unchanged_count = 0
    tickets_updated_count = 0
    for update in updates:
        row = rows_by_id.get(update.id)
        if row is None:
            continue
        next_functional_track = text_or_none(update.functional_track)
        next_is_in_scope = bool(update.is_in_scope)
        previous_functional_track = row.functional_track
        previous_is_in_scope = bool(row.is_in_scope)
        if (
            previous_functional_track == next_functional_track
            and previous_is_in_scope == next_is_in_scope
        ):
            unchanged_count += 1
            continue

        row.functional_track = next_functional_track
        row.is_in_scope = next_is_in_scope
        tickets_updated = update_tickets_for_scope_change(
            db,
            project_id=project_id,
            assignment_group_key=row.assignment_group_key,
            functional_track=next_functional_track,
            is_in_scope=next_is_in_scope,
        )
        tickets_updated_count += tickets_updated
        changes.append(
            AssignmentGroupScopeChange(
                id=row.id,
                assignment_group=row.assignment_group,
                previous_functional_track=previous_functional_track,
                next_functional_track=next_functional_track,
                previous_is_in_scope=previous_is_in_scope,
                next_is_in_scope=next_is_in_scope,
                tickets_updated=tickets_updated,
            )
        )

    return AssignmentGroupScopeUpdateResult(
        project_id=project_id,
        submitted_count=submitted_count,
        changed_count=len(changes),
        unchanged_count=unchanged_count,
        tickets_updated_count=tickets_updated_count,
        missing_count=len(missing_ids),
        changes=changes,
        warnings=warnings,
    )
