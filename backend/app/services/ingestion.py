from __future__ import annotations

import csv
import hashlib
import json
import logging
import re
from collections import Counter
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models import IngestionJob, TicketRawRow, UploadBatch, UploadedFile
from app.services.upload_lifecycle import (
    BATCH_STATUS_ARCHIVED,
    BATCH_STATUS_DELETED,
    BATCH_STATUS_INGESTED,
    BATCH_STATUS_INGESTING,
    BATCH_STATUS_INGESTION_FAILED,
    BATCH_STATUS_NORMALIZATION_FAILED,
    BATCH_STATUS_NORMALIZED,
    BATCH_STATUS_NORMALIZING,
    BATCH_STATUS_UPLOADED,
)

INGESTION_BATCH_SIZE = 1_000
CSV_ENCODING_CANDIDATES = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
LIKELY_TICKET_ID_COLUMNS = (
    "number",
    "ticket_id",
    "incident_number",
    "request_number",
    "task_number",
    "sc_task",
    "sctask",
    "sys_id",
)
logger = logging.getLogger(__name__)
LIKELY_CREATED_DATE_COLUMNS = (
    "sys_created_on",
    "created_at",
    "opened_at",
    "opened",
    "created",
    "created_date",
    "opened_date",
)


class IngestionError(Exception):
    pass


class UnsupportedIngestionFileTypeError(IngestionError):
    pass


class CsvEncodingDetectionError(IngestionError):
    pass


class IngestionJobAlreadyRunningError(IngestionError):
    pass


@dataclass(frozen=True)
class ParsedSourceRow:
    row_number: int
    raw_data: dict[str, Any]


@dataclass(frozen=True)
class FileRowCount:
    uploaded_file_id: UUID
    original_filename: str
    saved_filename: str | None
    row_count: int


@dataclass(frozen=True)
class ValidationSummary:
    upload_batch_id: UUID
    total_raw_rows: int
    missing_ticket_id_count: int
    missing_created_date_count: int
    duplicate_ticket_id_count: int
    duplicate_ticket_ids: dict[str, int]
    detected_source_columns: list[str]
    rows_by_uploaded_file: list[FileRowCount]
    message: str | None = None


def normalize_source_column_name(column_name: str) -> str:
    normalized = column_name.strip().lower()
    normalized = re.sub(r"[\s\-]+", "_", normalized)
    return re.sub(r"[^a-z0-9_]", "", normalized)


def make_unique_headers(headers: list[Any]) -> list[str]:
    used_headers: Counter[str] = Counter()
    unique_headers: list[str] = []

    for index, header in enumerate(headers, start=1):
        header_text = str(header).strip() if header is not None else ""
        header_text = header_text or f"column_{index}"
        used_headers[header_text] += 1
        if used_headers[header_text] > 1:
            header_text = f"{header_text}__{used_headers[header_text]}"
        unique_headers.append(header_text)

    return unique_headers


def to_json_value(value: Any) -> Any:
    if isinstance(value, datetime | date):
        return value.isoformat()

    if isinstance(value, Decimal):
        return float(value)

    if value is None or isinstance(value, str | int | float | bool):
        return value

    return str(value)


def build_raw_data(headers: list[str], values: tuple[Any, ...] | list[Any]) -> dict[str, Any]:
    raw_data = {
        header: to_json_value(values[index]) if index < len(values) else None
        for index, header in enumerate(headers)
    }

    if len(values) > len(headers):
        for extra_index, value in enumerate(values[len(headers) :], start=1):
            raw_data[f"__extra_{extra_index}"] = to_json_value(value)

    return raw_data


def row_has_any_value(raw_data: dict[str, Any]) -> bool:
    return any(value is not None and str(value).strip() for value in raw_data.values())


def find_likely_value(raw_data: dict[str, Any], likely_columns: tuple[str, ...]) -> str | None:
    normalized_lookup = {
        normalize_source_column_name(column_name): value for column_name, value in raw_data.items()
    }

    for likely_column in likely_columns:
        value = normalized_lookup.get(likely_column)
        if value is not None and str(value).strip():
            return str(value).strip()

    return None


def find_likely_ticket_id(raw_data: dict[str, Any]) -> str | None:
    return find_likely_value(raw_data, LIKELY_TICKET_ID_COLUMNS)


def find_likely_created_date(raw_data: dict[str, Any]) -> str | None:
    return find_likely_value(raw_data, LIKELY_CREATED_DATE_COLUMNS)


def hash_raw_row(raw_data: dict[str, Any]) -> str:
    payload = json.dumps(raw_data, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def validate_csv_encoding(path: Path, encoding: str) -> None:
    with path.open("r", encoding=encoding, newline="") as csv_file:
        reader = csv.reader(csv_file)
        for _ in reader:
            pass


def detect_csv_encoding(path: Path) -> str:
    failures: list[str] = []
    for encoding in CSV_ENCODING_CANDIDATES:
        try:
            validate_csv_encoding(path, encoding)
        except UnicodeDecodeError as exc:
            failures.append(f"{encoding}: {exc}")
            logger.info("CSV encoding %s failed for %s: %s", encoding, path.name, exc)
            continue

        logger.info("Using CSV encoding %s for %s", encoding, path.name)
        return encoding

    raise CsvEncodingDetectionError(
        "Unable to decode CSV using supported encodings: " + "; ".join(failures)
    )


def iter_csv_rows(path: Path) -> Iterator[ParsedSourceRow]:
    encoding = detect_csv_encoding(path)
    with path.open("r", encoding=encoding, newline="") as csv_file:
        reader = csv.reader(csv_file)
        try:
            headers = make_unique_headers(next(reader))
        except StopIteration:
            return

        for row_number, values in enumerate(reader, start=2):
            raw_data = build_raw_data(headers, values)
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(row_number=row_number, raw_data=raw_data)


def iter_xlsx_rows(path: Path) -> Iterator[ParsedSourceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.worksheets[0]
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = make_unique_headers(list(next(rows)))
        except StopIteration:
            return

        for row_number, values in enumerate(rows, start=2):
            raw_data = build_raw_data(headers, list(values))
            if row_has_any_value(raw_data):
                yield ParsedSourceRow(row_number=row_number, raw_data=raw_data)
    finally:
        workbook.close()


def iter_ticket_file_rows(path: Path) -> Iterator[ParsedSourceRow]:
    extension = path.suffix.lower()
    if extension == ".csv":
        yield from iter_csv_rows(path)
        return

    if extension == ".xlsx":
        yield from iter_xlsx_rows(path)
        return

    raise UnsupportedIngestionFileTypeError(f"Unsupported file extension: {extension}")


def get_ingestion_job_for_file(db: Session, uploaded_file_id: UUID) -> IngestionJob | None:
    statement = (
        select(IngestionJob)
        .where(IngestionJob.uploaded_file_id == uploaded_file_id)
        .order_by(IngestionJob.created_at.desc())
        .limit(1)
    )
    return db.scalar(statement)


def get_or_create_ingestion_job(db: Session, uploaded_file: UploadedFile) -> IngestionJob:
    ingestion_job = get_ingestion_job_for_file(db, uploaded_file.id)
    if ingestion_job is not None:
        return ingestion_job

    ingestion_job = IngestionJob(
        upload_batch_id=uploaded_file.upload_batch_id,
        uploaded_file_id=uploaded_file.id,
        job_type="FILE_INGESTION",
        status="PENDING",
        rows_total=0,
        rows_processed=0,
    )
    db.add(ingestion_job)
    db.flush()
    return ingestion_job


def recalculate_upload_batch_status(db: Session, upload_batch_id: UUID) -> UploadBatch | None:
    db.flush()
    upload_batch = db.get(UploadBatch, upload_batch_id)
    if upload_batch is None:
        return None
    if upload_batch.status in {
        BATCH_STATUS_NORMALIZING,
        BATCH_STATUS_NORMALIZED,
        BATCH_STATUS_NORMALIZATION_FAILED,
        BATCH_STATUS_ARCHIVED,
        BATCH_STATUS_DELETED,
    }:
        return upload_batch

    uploaded_file_count = db.scalar(
        select(func.count(UploadedFile.id)).where(UploadedFile.upload_batch_id == upload_batch_id)
    )
    uploaded_file_count = int(uploaded_file_count or 0)

    file_job_rows = list(
        db.execute(
            select(
                IngestionJob.uploaded_file_id,
                IngestionJob.status,
                IngestionJob.error_message,
            ).where(IngestionJob.upload_batch_id == upload_batch_id)
        )
    )
    job_statuses = [status for _, status, _ in file_job_rows]

    for uploaded_file_id, job_status, error_message in file_job_rows:
        if uploaded_file_id is None:
            continue
        uploaded_file = db.get(UploadedFile, uploaded_file_id)
        if uploaded_file is None:
            continue
        if job_status == "COMPLETED":
            uploaded_file.status = "INGESTED"
            uploaded_file.error_message = None
        elif job_status == "RUNNING":
            uploaded_file.status = "INGESTING"
            uploaded_file.error_message = None
        elif job_status == "FAILED":
            uploaded_file.status = "FAILED"
            uploaded_file.error_message = error_message

    if uploaded_file_count == 0 or not job_statuses:
        upload_batch.status = BATCH_STATUS_UPLOADED
        return upload_batch

    status_counts = Counter(job_statuses)
    attempted_count = status_counts["COMPLETED"] + status_counts["FAILED"]

    if status_counts["RUNNING"] > 0:
        upload_batch.status = BATCH_STATUS_INGESTING
        upload_batch.completed_at = None
        if upload_batch.started_at is None:
            upload_batch.started_at = datetime.now(UTC)
    elif status_counts["COMPLETED"] == uploaded_file_count:
        upload_batch.status = BATCH_STATUS_INGESTED
        upload_batch.completed_at = datetime.now(UTC)
    elif status_counts["FAILED"] == uploaded_file_count:
        upload_batch.status = BATCH_STATUS_INGESTION_FAILED
        upload_batch.completed_at = datetime.now(UTC)
    elif status_counts["FAILED"] > 0 and attempted_count > 0:
        upload_batch.status = BATCH_STATUS_INGESTION_FAILED
        upload_batch.completed_at = None
    elif attempted_count > 0:
        upload_batch.status = BATCH_STATUS_INGESTING
        upload_batch.completed_at = None
    else:
        upload_batch.status = BATCH_STATUS_UPLOADED
        upload_batch.completed_at = None

    return upload_batch


def ingest_uploaded_file(db: Session, uploaded_file_id: UUID) -> IngestionJob:
    uploaded_file = db.get(UploadedFile, uploaded_file_id)
    if uploaded_file is None:
        raise FileNotFoundError(f"Uploaded file {uploaded_file_id} was not found.")

    file_path = Path(uploaded_file.storage_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Stored file was not found: {file_path}")

    ingestion_job = get_or_create_ingestion_job(db, uploaded_file)
    if ingestion_job.status == "RUNNING":
        raise IngestionJobAlreadyRunningError(
            f"Ingestion job {ingestion_job.id} is already running."
        )

    now = datetime.now(UTC)
    ingestion_job.status = "RUNNING"
    ingestion_job.started_at = now
    ingestion_job.completed_at = None
    ingestion_job.error_message = None
    ingestion_job.rows_total = 0
    ingestion_job.rows_processed = 0
    uploaded_file.status = "INGESTING"
    uploaded_file.error_message = None
    recalculate_upload_batch_status(db, uploaded_file.upload_batch_id)
    db.commit()
    db.refresh(ingestion_job)

    db.execute(delete(TicketRawRow).where(TicketRawRow.uploaded_file_id == uploaded_file.id))
    db.commit()

    processed_row_count = 0
    pending_rows: list[TicketRawRow] = []

    try:
        for parsed_row in iter_ticket_file_rows(file_path):
            raw_ticket_number = find_likely_ticket_id(parsed_row.raw_data)
            pending_rows.append(
                TicketRawRow(
                    project_id=uploaded_file.project_id,
                    upload_batch_id=uploaded_file.upload_batch_id,
                    uploaded_file_id=uploaded_file.id,
                    ticket_type=uploaded_file.ticket_type,
                    row_number=parsed_row.row_number,
                    source_filename=uploaded_file.original_filename,
                    raw_ticket_number=raw_ticket_number,
                    raw_data=parsed_row.raw_data,
                    row_hash=hash_raw_row(parsed_row.raw_data),
                )
            )
            processed_row_count += 1

            if len(pending_rows) >= INGESTION_BATCH_SIZE:
                db.add_all(pending_rows)
                ingestion_job.rows_processed = processed_row_count
                ingestion_job.rows_total = processed_row_count
                db.commit()
                pending_rows.clear()

        if pending_rows:
            db.add_all(pending_rows)

        ingestion_job.status = "COMPLETED"
        ingestion_job.rows_processed = processed_row_count
        ingestion_job.rows_total = processed_row_count
        ingestion_job.completed_at = datetime.now(UTC)
        uploaded_file.status = "INGESTED"
        uploaded_file.error_message = None
        recalculate_upload_batch_status(db, uploaded_file.upload_batch_id)
        db.commit()
    except Exception as exc:
        db.rollback()
        db.execute(delete(TicketRawRow).where(TicketRawRow.uploaded_file_id == uploaded_file.id))
        ingestion_job = db.get(IngestionJob, ingestion_job.id)
        uploaded_file = db.get(UploadedFile, uploaded_file.id)
        if ingestion_job is not None:
            ingestion_job.status = "FAILED"
            ingestion_job.rows_processed = processed_row_count
            ingestion_job.rows_total = processed_row_count
            ingestion_job.error_message = str(exc)
            ingestion_job.completed_at = datetime.now(UTC)
        if uploaded_file is not None:
            uploaded_file.status = "FAILED"
            uploaded_file.error_message = str(exc)
            recalculate_upload_batch_status(db, uploaded_file.upload_batch_id)
        db.commit()
        if isinstance(exc, UnicodeDecodeError):
            raise IngestionError(f"CSV decoding failed: {exc}") from exc
        if isinstance(exc, IngestionError | OSError | SQLAlchemyError):
            raise
        raise IngestionError(str(exc)) from exc

    db.refresh(ingestion_job)
    return ingestion_job


def build_validation_summary(db: Session, upload_batch_id: UUID) -> ValidationSummary:
    upload_batch = db.get(UploadBatch, upload_batch_id)
    if upload_batch is None:
        raise FileNotFoundError(f"Upload batch {upload_batch_id} was not found.")

    total_raw_rows = 0
    missing_ticket_id_count = 0
    missing_created_date_count = 0
    detected_source_columns: set[str] = set()

    row_statement = (
        select(TicketRawRow)
        .where(TicketRawRow.upload_batch_id == upload_batch_id)
        .order_by(TicketRawRow.uploaded_file_id.asc(), TicketRawRow.row_number.asc())
    )
    for raw_row in db.scalars(row_statement).yield_per(INGESTION_BATCH_SIZE):
        total_raw_rows += 1
        detected_source_columns.update(raw_row.raw_data.keys())
        if not find_likely_ticket_id(raw_row.raw_data):
            missing_ticket_id_count += 1
        if not find_likely_created_date(raw_row.raw_data):
            missing_created_date_count += 1

    duplicate_statement = (
        select(TicketRawRow.raw_ticket_number, func.count(TicketRawRow.id))
        .where(
            TicketRawRow.upload_batch_id == upload_batch_id,
            TicketRawRow.raw_ticket_number.is_not(None),
            TicketRawRow.raw_ticket_number != "",
        )
        .group_by(TicketRawRow.raw_ticket_number)
        .having(func.count(TicketRawRow.id) > 1)
        .order_by(func.count(TicketRawRow.id).desc())
        .limit(100)
    )
    duplicate_ticket_ids = {
        str(ticket_id): int(count) for ticket_id, count in db.execute(duplicate_statement)
    }
    duplicate_count_statement = (
        select(func.count())
        .select_from(
            select(TicketRawRow.raw_ticket_number)
            .where(
                TicketRawRow.upload_batch_id == upload_batch_id,
                TicketRawRow.raw_ticket_number.is_not(None),
                TicketRawRow.raw_ticket_number != "",
            )
            .group_by(TicketRawRow.raw_ticket_number)
            .having(func.count(TicketRawRow.id) > 1)
            .subquery()
        )
    )
    duplicate_ticket_id_count = int(db.scalar(duplicate_count_statement) or 0)

    rows_by_file_statement = (
        select(
            UploadedFile.id,
            UploadedFile.original_filename,
            UploadedFile.saved_filename,
            func.count(TicketRawRow.id),
        )
        .outerjoin(TicketRawRow, TicketRawRow.uploaded_file_id == UploadedFile.id)
        .where(UploadedFile.upload_batch_id == upload_batch_id)
        .group_by(
            UploadedFile.id,
            UploadedFile.original_filename,
            UploadedFile.saved_filename,
            UploadedFile.created_at,
        )
        .order_by(UploadedFile.created_at.asc())
    )
    rows_by_uploaded_file = [
        FileRowCount(
            uploaded_file_id=uploaded_file_id,
            original_filename=original_filename,
            saved_filename=saved_filename,
            row_count=int(row_count),
        )
        for uploaded_file_id, original_filename, saved_filename, row_count in db.execute(
            rows_by_file_statement
        )
    ]

    return ValidationSummary(
        upload_batch_id=upload_batch_id,
        total_raw_rows=total_raw_rows,
        missing_ticket_id_count=missing_ticket_id_count,
        missing_created_date_count=missing_created_date_count,
        duplicate_ticket_id_count=duplicate_ticket_id_count,
        duplicate_ticket_ids=duplicate_ticket_ids,
        detected_source_columns=sorted(detected_source_columns),
        rows_by_uploaded_file=rows_by_uploaded_file,
        message="No raw rows found. Ingest files first." if total_raw_rows == 0 else None,
    )
