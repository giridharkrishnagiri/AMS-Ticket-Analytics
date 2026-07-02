from __future__ import annotations

import csv
from collections.abc import Iterator
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import Session

from app.models import AssignmentGroupMasterReference, Project
from app.services.in_scope_assignment_groups import normalize_assignment_group_key, text_or_none
from app.services.ingestion import (
    CSV_ENCODING_CANDIDATES,
    ParsedSourceRow,
    build_raw_data,
    detect_csv_encoding,
    make_unique_headers,
    normalize_source_column_name,
    row_has_any_value,
)

MAX_MESSAGE_SAMPLES = 50
PREFERRED_SHEET_NAME = "Master"

ASSIGNMENT_GROUP_ALIASES = (
    "Name",
    "Assignment Group",
    "Assignment group",
    "Assignment Groups",
    "Assignment group name",
    "Group",
    "Support Group",
)
DESCRIPTION_ALIASES = (
    "Description",
    "Desc",
    "Group Description",
    "Assignment Group Description",
)
MANAGER_ALIASES = (
    "Manager",
    "Manager Name",
    "Group Manager",
    "Assignment Group Manager",
    "Owner",
)


class AssignmentGroupMasterReferenceError(Exception):
    pass


@dataclass(frozen=True)
class AssignmentGroupMasterPreviewRow:
    assignment_group: str
    description: str | None
    manager_name: str | None
    source_sheet_name: str | None
    source_row_number: int | None


@dataclass
class AssignmentGroupMasterImportResult:
    project_id: UUID
    source_filename: str
    total_rows: int = 0
    imported_count: int = 0
    manager_populated_count: int = 0
    skipped_count: int = 0
    duplicate_count: int = 0
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    preview_rows: list[AssignmentGroupMasterPreviewRow] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return len(self.errors)

    @property
    def warning_count(self) -> int:
        return len(self.warnings)


@dataclass(frozen=True)
class AssignmentGroupMasterStatus:
    project_id: UUID
    active_count: int
    manager_populated_count: int
    last_imported_at: Any | None
    last_imported_filename: str | None
    preview_rows: list[AssignmentGroupMasterPreviewRow]


@dataclass(frozen=True)
class ParsedMasterReferenceRow:
    source_sheet_name: str | None
    row: ParsedSourceRow


def append_sample_message(messages: list[str], message: str) -> None:
    if len(messages) < MAX_MESSAGE_SAMPLES:
        messages.append(message)


def header_lookup(raw_data: dict[str, Any]) -> dict[str, Any]:
    return {
        normalize_source_column_name(column_name): value
        for column_name, value in raw_data.items()
    }


def get_aliased_value(raw_data: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    for alias in aliases:
        if alias in raw_data:
            return raw_data[alias]
    lookup = header_lookup(raw_data)
    for alias in aliases:
        normalized_alias = normalize_source_column_name(alias)
        if normalized_alias in lookup:
            return lookup[normalized_alias]
    return None


def has_alias(headers: set[str], aliases: tuple[str, ...]) -> bool:
    return any(normalize_source_column_name(alias) in headers for alias in aliases)


def validate_headers(raw_data: dict[str, Any]) -> None:
    headers = {normalize_source_column_name(column_name) for column_name in raw_data}
    if not has_alias(headers, ASSIGNMENT_GROUP_ALIASES):
        raise AssignmentGroupMasterReferenceError(
            "Assignment Group Master Reference workbook is missing required column: "
            "Name / Assignment Group.",
        )


def iter_master_csv_rows(path: Path) -> Iterator[ParsedMasterReferenceRow]:
    encoding = detect_csv_encoding(path)
    with path.open("r", encoding=encoding, newline="") as csv_file:
        reader = csv.reader(csv_file)
        try:
            headers = make_unique_headers(next(reader))
        except StopIteration:
            return

        for row_number, values in enumerate(reader, start=2):
            raw_data = build_raw_data(headers, values)
            if row_has_any_value(raw_data):
                yield ParsedMasterReferenceRow(
                    source_sheet_name="CSV",
                    row=ParsedSourceRow(row_number=row_number, raw_data=raw_data),
                )


def iter_master_xlsx_rows(
    path: Path,
    result: AssignmentGroupMasterImportResult,
) -> Iterator[ParsedMasterReferenceRow]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = (
            workbook[PREFERRED_SHEET_NAME]
            if PREFERRED_SHEET_NAME in workbook.sheetnames
            else workbook.worksheets[0]
        )
        if worksheet.title != PREFERRED_SHEET_NAME:
            append_sample_message(
                result.warnings,
                f"Sheet '{PREFERRED_SHEET_NAME}' was not found; imported first sheet "
                f"'{worksheet.title}'.",
            )
        rows = worksheet.iter_rows(values_only=True)
        try:
            headers = make_unique_headers(next(rows))
        except StopIteration:
            return
        for row_number, values in enumerate(rows, start=2):
            raw_data = build_raw_data(headers, list(values))
            if row_has_any_value(raw_data):
                yield ParsedMasterReferenceRow(
                    source_sheet_name=worksheet.title,
                    row=ParsedSourceRow(row_number=row_number, raw_data=raw_data),
                )
    finally:
        workbook.close()


def iter_master_reference_rows(
    path: Path,
    result: AssignmentGroupMasterImportResult,
) -> Iterator[ParsedMasterReferenceRow]:
    extension = path.suffix.lower()
    if extension == ".csv":
        yield from iter_master_csv_rows(path)
        return
    if extension == ".xlsx":
        yield from iter_master_xlsx_rows(path, result)
        return
    raise AssignmentGroupMasterReferenceError(
        f"Unsupported assignment group master reference extension: {extension}. Use XLSX or CSV.",
    )


def ensure_project(db: Session, project_id: UUID) -> Project:
    project = db.get(Project, project_id)
    if project is None:
        raise FileNotFoundError(f"Project {project_id} was not found.")
    return project


def import_assignment_group_master_reference(
    db: Session,
    project_id: UUID,
    path: Path,
    source_filename: str,
) -> AssignmentGroupMasterImportResult:
    project = ensure_project(db, project_id)
    result = AssignmentGroupMasterImportResult(
        project_id=project_id,
        source_filename=source_filename,
    )
    rows_by_key: dict[str, dict[str, Any]] = {}

    try:
        first_row_checked = False
        for parsed in iter_master_reference_rows(path, result):
            parsed_row = parsed.row
            result.total_rows += 1
            if not first_row_checked:
                validate_headers(parsed_row.raw_data)
                first_row_checked = True

            assignment_group = text_or_none(
                get_aliased_value(parsed_row.raw_data, ASSIGNMENT_GROUP_ALIASES)
            )
            if assignment_group is None:
                result.skipped_count += 1
                continue

            assignment_group_key = normalize_assignment_group_key(assignment_group)
            if assignment_group_key is None:
                result.skipped_count += 1
                continue

            description = text_or_none(
                get_aliased_value(parsed_row.raw_data, DESCRIPTION_ALIASES)
            )
            manager_name = text_or_none(get_aliased_value(parsed_row.raw_data, MANAGER_ALIASES))
            if assignment_group_key in rows_by_key:
                result.duplicate_count += 1
            rows_by_key[assignment_group_key] = {
                "client_id": project.client_id,
                "project_id": project_id,
                "assignment_group": assignment_group,
                "assignment_group_key": assignment_group_key,
                "description": description,
                "manager_name": manager_name,
                "source_filename": source_filename,
                "source_sheet_name": parsed.source_sheet_name,
                "source_row_number": parsed_row.row_number,
                "is_active": True,
            }

        if result.total_rows == 0:
            raise AssignmentGroupMasterReferenceError(
                "Assignment group master reference file did not contain any data rows.",
            )

        db.execute(
            delete(AssignmentGroupMasterReference).where(
                AssignmentGroupMasterReference.project_id == project_id,
            )
        )
        for values in rows_by_key.values():
            db.add(AssignmentGroupMasterReference(**values))

        result.imported_count = len(rows_by_key)
        result.manager_populated_count = sum(
            1 for values in rows_by_key.values() if values.get("manager_name")
        )
        result.preview_rows = [
            AssignmentGroupMasterPreviewRow(
                assignment_group=str(values["assignment_group"]),
                description=values["description"],
                manager_name=values["manager_name"],
                source_sheet_name=values["source_sheet_name"],
                source_row_number=values["source_row_number"],
            )
            for values in list(rows_by_key.values())[:25]
        ]
        db.flush()
    except SQLAlchemyError as exc:
        db.rollback()
        raise AssignmentGroupMasterReferenceError(
            f"Assignment group master reference could not be saved: {exc}",
        ) from exc
    except UnicodeDecodeError as exc:
        db.rollback()
        raise AssignmentGroupMasterReferenceError(
            "Assignment group master reference CSV could not be decoded using supported "
            f"encodings: {', '.join(CSV_ENCODING_CANDIDATES)}",
        ) from exc
    except Exception:
        db.rollback()
        raise

    return result


def assignment_group_master_reference_status(
    db: Session,
    project_id: UUID,
    *,
    preview_limit: int = 25,
) -> AssignmentGroupMasterStatus:
    ensure_project(db, project_id)
    active_count = int(
        db.scalar(
            select(func.count(AssignmentGroupMasterReference.id)).where(
                AssignmentGroupMasterReference.project_id == project_id,
                AssignmentGroupMasterReference.is_active.is_(True),
            )
        )
        or 0
    )
    manager_populated_count = int(
        db.scalar(
            select(func.count(AssignmentGroupMasterReference.id)).where(
                AssignmentGroupMasterReference.project_id == project_id,
                AssignmentGroupMasterReference.is_active.is_(True),
                AssignmentGroupMasterReference.manager_name.is_not(None),
                func.btrim(AssignmentGroupMasterReference.manager_name) != "",
            )
        )
        or 0
    )
    last_imported_at = db.scalar(
        select(func.max(AssignmentGroupMasterReference.created_at)).where(
            AssignmentGroupMasterReference.project_id == project_id,
            AssignmentGroupMasterReference.is_active.is_(True),
        )
    )
    last_imported_filename = db.scalar(
        select(AssignmentGroupMasterReference.source_filename)
        .where(
            AssignmentGroupMasterReference.project_id == project_id,
            AssignmentGroupMasterReference.is_active.is_(True),
        )
        .order_by(
            AssignmentGroupMasterReference.created_at.desc(),
            AssignmentGroupMasterReference.id.desc(),
        )
        .limit(1)
    )
    preview = db.scalars(
        select(AssignmentGroupMasterReference)
        .where(
            AssignmentGroupMasterReference.project_id == project_id,
            AssignmentGroupMasterReference.is_active.is_(True),
        )
        .order_by(
            AssignmentGroupMasterReference.assignment_group.asc(),
            AssignmentGroupMasterReference.id.asc(),
        )
        .limit(preview_limit),
    ).all()
    return AssignmentGroupMasterStatus(
        project_id=project_id,
        active_count=active_count,
        manager_populated_count=manager_populated_count,
        last_imported_at=last_imported_at,
        last_imported_filename=last_imported_filename,
        preview_rows=[
            AssignmentGroupMasterPreviewRow(
                assignment_group=row.assignment_group,
                description=row.description,
                manager_name=row.manager_name,
                source_sheet_name=row.source_sheet_name,
                source_row_number=row.source_row_number,
            )
            for row in preview
        ],
    )


def list_assignment_group_master_references(
    db: Session,
    project_id: UUID,
    *,
    limit: int = 100,
    offset: int = 0,
) -> list[AssignmentGroupMasterReference]:
    ensure_project(db, project_id)
    return list(
        db.scalars(
            select(AssignmentGroupMasterReference)
            .where(
                AssignmentGroupMasterReference.project_id == project_id,
                AssignmentGroupMasterReference.is_active.is_(True),
            )
            .order_by(
                AssignmentGroupMasterReference.assignment_group.asc(),
                AssignmentGroupMasterReference.id.asc(),
            )
            .offset(offset)
            .limit(limit),
        ).all()
    )


def active_assignment_group_master_manager_map(
    db: Session,
    project_id: UUID,
) -> dict[str, str]:
    ensure_project(db, project_id)
    rows = db.execute(
        select(
            AssignmentGroupMasterReference.assignment_group_key,
            AssignmentGroupMasterReference.manager_name,
        ).where(
            AssignmentGroupMasterReference.project_id == project_id,
            AssignmentGroupMasterReference.is_active.is_(True),
            AssignmentGroupMasterReference.manager_name.is_not(None),
            func.btrim(AssignmentGroupMasterReference.manager_name) != "",
        )
    ).all()
    return {
        str(assignment_group_key): str(manager_name).strip()
        for assignment_group_key, manager_name in rows
        if assignment_group_key and manager_name and str(manager_name).strip()
    }
