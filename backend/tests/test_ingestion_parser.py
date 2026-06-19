import csv
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import delete, func, select

from app.db.session import SessionLocal
from app.models import Client, IngestionJob, Project, TicketRawRow, UploadBatch, UploadedFile
from app.services.ingestion import (
    IngestionError,
    detect_csv_encoding,
    find_likely_created_date,
    find_likely_ticket_id,
    hash_raw_row,
    ingest_uploaded_file,
    iter_ticket_file_rows,
)


def test_utf8_sig_csv_parser_streams_rows(tmp_path) -> None:
    csv_path = tmp_path / "incidents.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(["number", "opened_at", "short_description"])
        writer.writerow(["INC001", "2026-06-01 09:00:00", "Email issue"])
        writer.writerow(["INC002", "", "Laptop issue"])

    rows = list(iter_ticket_file_rows(csv_path))

    assert [row.row_number for row in rows] == [2, 3]
    assert rows[0].raw_data["number"] == "INC001"
    assert find_likely_ticket_id(rows[0].raw_data) == "INC001"
    assert find_likely_created_date(rows[0].raw_data) == "2026-06-01 09:00:00"
    assert find_likely_created_date(rows[1].raw_data) is None


def test_cp1252_csv_parser_retries_after_utf8_failure(tmp_path) -> None:
    csv_path = tmp_path / "incidents_cp1252.csv"
    csv_path.write_bytes(
        "number,opened_at,short_description\r\n"
        "INC003,2026-06-03 10:00:00,Printer issue ™\r\n".encode("cp1252")
    )

    rows = list(iter_ticket_file_rows(csv_path))

    assert detect_csv_encoding(csv_path) == "cp1252"
    assert len(rows) == 1
    assert rows[0].raw_data["number"] == "INC003"
    assert rows[0].raw_data["short_description"] == "Printer issue ™"


def test_xlsx_parser_streams_first_worksheet(tmp_path) -> None:
    xlsx_path = tmp_path / "tasks.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Tasks"
    worksheet.append(["task_number", "created", "state"])
    worksheet.append(["TASK001", "2026-06-02", "Open"])
    workbook.create_sheet("Ignored").append(["task_number"])
    workbook.save(xlsx_path)

    rows = list(iter_ticket_file_rows(xlsx_path))

    assert len(rows) == 1
    assert rows[0].row_number == 2
    assert rows[0].raw_data["task_number"] == "TASK001"
    assert find_likely_ticket_id(rows[0].raw_data) == "TASK001"
    assert find_likely_created_date(rows[0].raw_data) == "2026-06-02"


def test_raw_row_hash_is_stable_for_same_payload() -> None:
    raw_data = {"number": "INC001", "opened_at": "2026-06-01"}

    assert hash_raw_row(raw_data) == hash_raw_row({"opened_at": "2026-06-01", "number": "INC001"})


def create_uploaded_file_fixture(file_path):
    db = SessionLocal()
    unique_suffix = uuid4().hex[:12]
    client = Client(
        name=f"Parser Test Client {unique_suffix}",
        code=f"PTC-{unique_suffix}",
    )
    db.add(client)
    db.flush()

    project = Project(
        client_id=client.id,
        name=f"Parser Test Project {unique_suffix}",
        code=f"PTP-{unique_suffix}",
    )
    db.add(project)
    db.flush()

    upload_batch = UploadBatch(
        project_id=project.id,
        month_key="2026-06",
        batch_name=f"Parser Test Batch {unique_suffix}",
        status="PENDING",
        file_count=1,
        total_size_bytes=file_path.stat().st_size,
    )
    db.add(upload_batch)
    db.flush()

    uploaded_file = UploadedFile(
        upload_batch_id=upload_batch.id,
        project_id=project.id,
        ticket_type="INCIDENT",
        original_filename=file_path.name,
        saved_filename=file_path.name,
        storage_path=str(file_path),
        size_bytes=file_path.stat().st_size,
        status="STORED",
    )
    db.add(uploaded_file)
    db.flush()

    ingestion_job = IngestionJob(
        upload_batch_id=upload_batch.id,
        uploaded_file_id=uploaded_file.id,
        job_type="FILE_INGESTION",
        status="PENDING",
        rows_total=0,
        rows_processed=0,
    )
    db.add(ingestion_job)
    db.commit()

    return db, client.id, uploaded_file.id, ingestion_job.id


def test_ingestion_failure_updates_job_status_and_error_message(tmp_path) -> None:
    bad_xlsx_path = tmp_path / "bad.xlsx"
    bad_xlsx_path.write_bytes(b"not a real workbook")
    db, client_id, uploaded_file_id, ingestion_job_id = create_uploaded_file_fixture(bad_xlsx_path)

    try:
        with pytest.raises(IngestionError):
            ingest_uploaded_file(db, uploaded_file_id)

        failed_job = db.get(IngestionJob, ingestion_job_id)
        assert failed_job is not None
        assert failed_job.status == "FAILED"
        assert failed_job.error_message
        assert failed_job.completed_at is not None
    finally:
        db.execute(delete(Client).where(Client.id == client_id))
        db.commit()
        db.close()


def test_retry_after_failure_does_not_duplicate_raw_rows(tmp_path) -> None:
    workbook_path = tmp_path / "retry.xlsx"
    workbook_path.write_bytes(b"not a real workbook")
    db, client_id, uploaded_file_id, _ = create_uploaded_file_fixture(workbook_path)

    try:
        with pytest.raises(IngestionError):
            ingest_uploaded_file(db, uploaded_file_id)

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["number", "opened_at"])
        worksheet.append(["INC004", "2026-06-04 11:00:00"])
        workbook.save(workbook_path)

        completed_job = ingest_uploaded_file(db, uploaded_file_id)
        assert completed_job.status == "COMPLETED"
        assert completed_job.processed_row_count == 1

        retried_job = ingest_uploaded_file(db, uploaded_file_id)
        assert retried_job.status == "COMPLETED"
        assert retried_job.processed_row_count == 1

        raw_row_count = db.scalar(
            select(func.count(TicketRawRow.id)).where(
                TicketRawRow.uploaded_file_id == uploaded_file_id
            )
        )
        assert raw_row_count == 1
    finally:
        db.execute(delete(Client).where(Client.id == client_id))
        db.commit()
        db.close()
